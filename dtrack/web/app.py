"""FastAPI web UI for dtrack — all routes in one file."""

import io
import json
import os
import sqlite3
import sys
from contextlib import redirect_stdout, redirect_stderr
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
import glob

from fastapi import FastAPI, Request, Query, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(title="dtrack")

_dir = Path(__file__).parent
app.mount("/static", StaticFiles(directory=_dir / "static"), name="static")
templates = Jinja2Templates(directory=_dir / "templates")

# These are set by serve() before uvicorn.run()
_db_path: str = ""
_db_dir: str = ""  # directory containing the database — all relative paths resolve from here
_config_path: str = ""
_original_config_path: str = ""  # saved when switching to testing mode
_testing_mode: bool = False


def _resolve(rel_path: str) -> str:
    """Resolve a path relative to the database directory."""
    p = os.path.join(_db_dir, rel_path)
    return os.path.normpath(p)


def _cfg():
    from ..config import load_unified_config
    return load_unified_config(_config_path)


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(request, "pairs.html")

@app.get("/pairs", response_class=HTMLResponse)
async def pairs_page(request: Request):
    return templates.TemplateResponse(request, "pairs.html")

@app.get("/load_row", response_class=HTMLResponse)
async def load_row_page(request: Request):
    return templates.TemplateResponse(request, "load_row.html")

@app.get("/row_compare", response_class=HTMLResponse)
async def row_compare_page(request: Request):
    return templates.TemplateResponse(request, "row_compare.html")

@app.get("/col_mapping", response_class=HTMLResponse)
async def col_mapping_page(request: Request):
    return templates.TemplateResponse(request, "col_mapping.html")

@app.get("/col_gen", response_class=HTMLResponse)
async def col_gen_page(request: Request):
    return templates.TemplateResponse(request, "col_gen.html")

@app.get("/load_col", response_class=HTMLResponse)
async def load_col_page(request: Request):
    return templates.TemplateResponse(request, "load_col.html")

@app.get("/col_compare", response_class=HTMLResponse)
async def col_compare_page(request: Request):
    return templates.TemplateResponse(request, "col_compare.html")

@app.get("/benchmark", response_class=HTMLResponse)
async def benchmark_page(request: Request):
    return templates.TemplateResponse(request, "benchmark.html")

@app.get("/playground", response_class=HTMLResponse)
async def playground_page(request: Request):
    return templates.TemplateResponse(request, "playground.html")

@app.get("/csv_compare", response_class=HTMLResponse)
async def csv_compare_page(request: Request):
    return templates.TemplateResponse(request, "csv_compare.html")


# ---------------------------------------------------------------------------
# API routes
# ---------------------------------------------------------------------------

@app.get("/api/status")
async def api_status():
    """Return pairs + metadata + date ranges."""
    from ..db import list_table_pairs, get_metadata, get_column_meta
    from ..config import load_unified_config, get_all_tables_from_unified

    pairs = list_table_pairs(_db_path)
    config = load_unified_config(_config_path)

    result = []
    for pair in pairs:
        pname = pair["pair_name"]
        meta_left = get_metadata(_db_path, pair["table_left"])
        meta_right = get_metadata(_db_path, pair["table_right"])
        cols_left = get_column_meta(_db_path, pair["table_left"])
        cols_right = get_column_meta(_db_path, pair["table_right"])

        pair_cfg = config.get("pairs", {}).get(pname, {})

        result.append({
            "pair_name": pname,
            "table_left": pair["table_left"],
            "table_right": pair["table_right"],
            "source_left": pair.get("source_left", ""),
            "source_right": pair.get("source_right", ""),
            "skip": pair_cfg.get("skip", False),
            "left": {
                "min_date": (meta_left or {}).get("min_date_loaded"),
                "max_date": (meta_left or {}).get("max_date_loaded"),
                "data_type": (meta_left or {}).get("data_type"),
                "row_count": (meta_left or {}).get("row_count_total"),
                "col_count": len(cols_left),
            },
            "right": {
                "min_date": (meta_right or {}).get("min_date_loaded"),
                "max_date": (meta_right or {}).get("max_date_loaded"),
                "data_type": (meta_right or {}).get("data_type"),
                "row_count": (meta_right or {}).get("row_count_total"),
                "col_count": len(cols_right),
            },
            "col_mappings": len(pair.get("col_mappings", {}) or {}),
        })

    return {"pairs": result, "db_path": _db_path, "config_path": _config_path}


@app.post("/api/init")
async def api_init():
    """Initialize or refresh the database."""
    from ..db import init_database, refresh_database, refresh_metadata_from_data

    if os.path.exists(_db_path):
        actions = refresh_database(_db_path)
        n_meta = refresh_metadata_from_data(_db_path)
        return {"action": "refreshed", "details": actions, "metadata_refreshed": n_meta}
    else:
        init_database(_db_path)
        return {"action": "created", "db_path": _db_path}


@app.post("/api/refresh-metadata")
async def api_refresh_metadata():
    """Recompute NULL metadata fields from actual data in _row_counts and _col_stats."""
    from ..db import refresh_metadata_from_data
    try:
        n = refresh_metadata_from_data(_db_path)
        return {"ok": True, "updated": n}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/date-types")
async def api_date_types():
    """Return supported date_type values and their SQL format examples."""
    from ..platforms.base import DATE_TYPE_FORMATS
    return {"date_types": DATE_TYPE_FORMATS}


@app.get("/api/constants")
async def api_constants():
    """Return UI constants defined once in dtrack/constants.py so the JS
    side doesn't need to keep a parallel copy."""
    from .. import constants as C
    return {
        "DATA_SOURCES": C.DATA_SOURCES,
        "CONNECTION_MACROS": C.CONNECTION_MACROS,
        "DATE_COLUMN_TYPES": C.DATE_COLUMN_TYPES,
    }


@app.post("/api/extract")
async def api_extract(request: Request):
    """Run gen-sas or gen-aws extraction."""
    body = await request.json()
    platform = body.get("platform", "aws")  # "sas" or "aws"
    extract_type = body.get("type", "row")  # "row" or "col"
    from_date = body.get("from_date")
    to_date = body.get("to_date")
    outdir = _resolve(body.get("outdir", "./csv/" if platform == "aws" else "./sas/"))
    max_workers = body.get("max_workers")

    buf = io.StringIO()
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            if platform == "aws":
                from ..platforms.athena import extract_aws
                extract_aws(_config_path, outdir, types=[extract_type],
                            db_path=_db_path, from_date=from_date, to_date=to_date,
                            max_workers=int(max_workers) if max_workers else None)
            else:
                from ..platforms.oracle import gen_sas
                gen_sas(_config_path, outdir, types=[extract_type],
                        db_path=_db_path, from_date=from_date, to_date=to_date)
        return {"ok": True, "output": buf.getvalue()}
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": str(e), "output": buf.getvalue()
        })


@app.post("/api/extract/run-sql")
async def api_run_sql_file(request: Request):
    """Execute queries from extract_row.sql via Athena.

    Streams progress as Server-Sent Events (SSE).  Each completed query
    emits a 'progress' event; the final message is a 'done' event with
    summary.  Progress also prints to the terminal (stderr).

    Body: {outdir, max_workers, type}
    """
    import queue
    import threading

    body = await request.json()
    extract_type = body.get("type", "row")
    outdir = _resolve(body.get("outdir", "./csv/"))
    max_workers = body.get("max_workers")
    resume = bool(body.get("resume", False))

    sql_path = os.path.join(outdir, f"extract_{extract_type}.sql")
    if not os.path.exists(sql_path):
        return JSONResponse(status_code=404, content={
            "ok": False, "error": f"SQL file not found: {sql_path}"
        })

    progress_q = queue.Queue()

    def _on_progress(result):
        progress_q.put(result)

    def _run():
        try:
            from ..platforms.athena import run_sql_file
            results = run_sql_file(
                sql_path, outdir,
                max_workers=int(max_workers) if max_workers else None,
                db_path=_db_path,
                on_progress=_on_progress,
                resume=resume,
            )
            ok_count = sum(1 for r in results if r.get("ok"))
            fail_count = len(results) - ok_count
            progress_q.put({"_done": True, "ok": fail_count == 0,
                            "results": results, "total": len(results),
                            "succeeded": ok_count, "failed": fail_count})
        except Exception as e:
            progress_q.put({"_done": True, "ok": False, "error": str(e)})

    async def _stream():
        thread = threading.Thread(target=_run, daemon=True)
        thread.start()
        while True:
            # Poll the queue (non-blocking in async context)
            import asyncio
            try:
                msg = progress_q.get_nowait()
            except queue.Empty:
                await asyncio.sleep(0.2)
                continue

            if msg.get("_done"):
                msg.pop("_done")
                yield f"event: done\ndata: {json.dumps(msg)}\n\n"
                break
            else:
                yield f"event: progress\ndata: {json.dumps(msg)}\n\n"

    return StreamingResponse(_stream(), media_type="text/event-stream")


def _check_pair_buckets(filtered_config, pair_names, from_date, to_date):
    """Emit a per-pair PASS/WARNING line confirming that left and right
    will be queried with the SAME vintage bucket boundaries.

    The log lines start with `[BUCKET CHECK]` so the col_gen UI can
    highlight them green (PASS) or amber (WARNING).
    """
    from ..date_utils import vintage_bucket_spans, bucket_date
    from ..db import get_row_comparison
    from ..platforms.base import qualified_name
    from ..config import _derive_side_name

    for pname in pair_names:
        pair_cfg = filtered_config.get("pairs", {}).get(pname)
        if not pair_cfg:
            continue
        vintage = pair_cfg.get("vintage", "") or ""
        if vintage in ("", "all"):
            print(f"[BUCKET CHECK] pair={pname} | vintage={vintage or 'none'} | "
                  f"single bucket on both sides | PASS")
            continue

        left_cfg = pair_cfg.get("left", {}).copy()
        right_cfg = pair_cfg.get("right", {}).copy()
        left_cfg["name"] = _derive_side_name(left_cfg, pname)
        right_cfg["name"] = _derive_side_name(right_cfg, pname)

        # Prefer real row-compare matching dates; fall back to from/to spans.
        comp = get_row_comparison(_db_path, pname)
        matching = (comp or {}).get("matching_dates") or []

        def _keys_for(cfg):
            if matching:
                return sorted({bucket_date(d, vintage) for d in matching})
            if from_date and to_date:
                return [k for k, _, _ in vintage_bucket_spans(from_date, to_date, vintage)]
            return []

        left_keys = _keys_for(left_cfg)
        right_keys = _keys_for(right_cfg)

        if not left_keys and not right_keys:
            print(f"[BUCKET CHECK] pair={pname} vintage={vintage} | "
                  f"no row-compare data and no from/to set — both sides will "
                  f"emit dt='all' | PASS")
            continue

        source = "row-compare matching dates" if matching else f"range {from_date}..{to_date}"
        if left_keys == right_keys:
            print(f"[BUCKET CHECK] pair={pname} vintage={vintage} | "
                  f"{len(left_keys)} buckets ({left_keys[0]} .. {left_keys[-1]}) "
                  f"identical on LEFT and RIGHT | source: {source} | PASS")
        else:
            # Find first divergence for the message
            diff_idx = next(
                (i for i, (a, b) in enumerate(zip(left_keys, right_keys)) if a != b),
                min(len(left_keys), len(right_keys)),
            )
            print(f"[BUCKET CHECK] pair={pname} vintage={vintage} | "
                  f"L={len(left_keys)} R={len(right_keys)} buckets | "
                  f"first divergence at index {diff_idx} | source: {source} | WARNING")


@app.post("/api/generate")
async def api_generate_files(request: Request):
    """Generate SAS and SQL files without running extraction.

    Writes extract_row.sas (for Oracle/SAS tables) and extract_row.sql
    (for AWS tables) to their respective output directories.
    """
    body = await request.json()
    extract_type = body.get("type", "row")
    from_date = body.get("from_date")
    to_date = body.get("to_date")
    sas_outdir = _resolve(body.get("sas_outdir", "./sas/"))
    aws_outdir = _resolve(body.get("aws_outdir", "./csv/"))
    pair_names = body.get("pair_names")  # None = all pairs

    buf = io.StringIO()
    results = {"sas_file": None, "sql_file": None}

    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            from ..config import load_unified_config
            from ..platforms.oracle import load_tables_from_config, inject_where_from_config

            config = load_unified_config(_config_path)

            # Filter config to selected pairs before loading tables
            if pair_names:
                filtered_config = dict(config)
                filtered_config["pairs"] = {
                    k: v for k, v in config.get("pairs", {}).items()
                    if k in pair_names
                }
            else:
                filtered_config = config

            # col_map and col_filter live in _table_pairs, not dtrack.json.
            # Hydrate them into the in-memory config so get_all_tables_from_unified
            # can resolve include/exclude patterns and produce _selected_cols.
            from ..db import get_pair_col_map_from_db, get_pair_col_filter
            for pname, pcfg in filtered_config.get("pairs", {}).items():
                if not pcfg.get("col_map"):
                    db_map = get_pair_col_map_from_db(_db_path, pname)
                    if db_map:
                        pcfg["col_map"] = db_map
                if not pcfg.get("col_filter"):
                    cf = get_pair_col_filter(_db_path, pname)
                    if cf.get("include") or cf.get("exclude"):
                        pcfg["col_filter"] = cf
                print(f"[hydrate] {pname}: "
                      f"col_map={len(pcfg.get('col_map') or {})} entries, "
                      f"col_filter={pcfg.get('col_filter') or 'NONE'}")

            all_tables = load_tables_from_config(filtered_config)
            inject_where_from_config(all_tables, filtered_config)

            # Inject excludeDates from pair config into table configs
            for tbl in all_tables:
                for pair_name in tbl.get('_pairs', []):
                    pair_cfg = filtered_config.get('pairs', {}).get(pair_name, {})
                    exc = pair_cfg.get('excludeDates', [])
                    if exc:
                        tbl['_exclude_dates'] = exc
                        break

            # Inject date bounds. Per-pair fromDate/toDate (set on the card)
            # overrides the request-level globals; when a pair has neither
            # its own nor a global value, the table skips the bound.
            for tbl in all_tables:
                date_col = tbl.get('date_col', '')
                if not date_col:
                    continue
                # Resolve per-pair dates from the first pair this table
                # belongs to. In multi-pair-per-table cases the first
                # pair's dates win (same behavior as col_filter dedup).
                pair_from, pair_to = None, None
                for pn in tbl.get('_pairs', []):
                    pc = filtered_config.get('pairs', {}).get(pn, {})
                    pair_from = pc.get('fromDate') or pair_from
                    pair_to = pc.get('toDate') or pair_to
                    if pair_from and pair_to:
                        break
                eff_from = pair_from or from_date
                eff_to   = pair_to   or to_date
                if not (eff_from or eff_to):
                    continue

                date_type = (tbl.get('date_type') or '').lower()
                source = tbl.get('source', '').lower()
                bounds = []
                if source == 'aws':
                    from ..platforms.athena import _format_athena_date_bound
                    if eff_from:
                        bounds.append(f"{date_col} >= {_format_athena_date_bound(eff_from, date_type, is_upper=False)}")
                    if eff_to:
                        bounds.append(f"{date_col} <= {_format_athena_date_bound(eff_to, date_type, is_upper=True)}")
                else:
                    from ..platforms.oracle import _format_date_bound, is_sas_table
                    is_sas = is_sas_table(tbl)
                    if eff_from:
                        bounds.append(f"{date_col} >= {_format_date_bound(eff_from, date_type, is_sas, is_upper=False)}")
                    if eff_to:
                        bounds.append(f"{date_col} <= {_format_date_bound(eff_to, date_type, is_sas, is_upper=True)}")
                if bounds:
                    extra = " AND ".join(bounds)
                    if source == 'aws':
                        # Stash separately so the AWS col path can skip them
                        # when its BETWEEN/IN filter already covers the range
                        # (prevents `col >= X AND col <= Y AND col BETWEEN X
                        # AND Y` duplication, matching the oracle fix).
                        tbl['_date_bounds'] = extra
                    else:
                        existing = tbl.get('where', '').strip()
                        tbl['where'] = f"({existing}) AND {extra}" if existing else extra
                # Stash the resolved from/to so col-stats can synthesize
                # vintage buckets when no row-compare data is loaded yet.
                tbl['_from_date'] = eff_from
                tbl['_to_date'] = eff_to
                print(f"[date bounds] {tbl['name']}: from={eff_from} to={eff_to} "
                      f"(source: {'pair' if (pair_from or pair_to) else 'global'})")

            # Generate SAS file — use filtered config if pairs selected
            oracle_tables = [t for t in all_tables
                             if t.get('source', '').lower() in ('oracle', 'sas', 'hadoop')]
            if oracle_tables:
                from ..platforms.oracle import gen_sas
                if pair_names:
                    import tempfile
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
                        json.dump(filtered_config, tmp, indent=2)
                        tmp_path = tmp.name
                    try:
                        gen_sas(tmp_path, sas_outdir, types=[extract_type],
                                db_path=_db_path, from_date=from_date, to_date=to_date)
                    finally:
                        os.unlink(tmp_path)
                else:
                    gen_sas(_config_path, sas_outdir, types=[extract_type],
                            db_path=_db_path, from_date=from_date, to_date=to_date)
                results["sas_file"] = os.path.join(sas_outdir, f"extract_{extract_type}.sas")

            # Generate SQL file
            aws_tables = [t for t in all_tables
                          if t.get('source', '').lower() == 'aws']
            if aws_tables:
                from ..platforms.athena import _write_combined_sql
                os.makedirs(aws_outdir, exist_ok=True)

                # Discover columns for AWS tables (needed for col stats SQL)
                from ..platforms.oracle import _discover_and_write_columns
                _discover_and_write_columns(aws_tables, aws_outdir, _db_path)

                # Fill columns from _column_meta into table configs for col SQL gen
                if extract_type == "col":
                    from ..platforms.base import fill_columns_from_meta
                    fill_columns_from_meta(aws_tables, _db_path)

                _write_combined_sql(aws_tables, aws_outdir, extract_type, db_path=_db_path)
                results["sql_file"] = os.path.join(aws_outdir, f"extract_{extract_type}.sql")

            # Cross-side bucket equality check (col extraction only — per-pair
            # vintage must produce the same bucket boundaries on left & right
            # so the downstream column comparison stays apples-to-apples).
            if extract_type == "col" and pair_names:
                _check_pair_buckets(filtered_config, pair_names, from_date, to_date)

        return {"ok": True, "output": buf.getvalue(), **results}
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": str(e), "output": buf.getvalue()
        })


@app.post("/api/load")
async def api_load(request: Request):
    """Run load-row or load-col."""
    body = await request.json()
    load_type = body.get("type", "row")  # "row" or "col"
    folder = _resolve(body.get("folder", "./csv/"))

    buf = io.StringIO()
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            from ..config import load_unified_config, get_all_tables_from_unified
            from ..platforms.base import qualified_name

            config = load_unified_config(_config_path)
            tables = get_all_tables_from_unified(config)

            if load_type == "row":
                from ..loader import load_row_counts
                from ..db import get_row_counts, insert_column_meta
                import csv as csv_mod
                for tbl in tables:
                    qname = qualified_name(tbl)
                    csv_path = os.path.join(folder, f"{qname}_row.csv")
                    if not os.path.exists(csv_path):
                        continue
                    load_row_counts(
                        db_path=_db_path, file_or_folder=csv_path,
                        table_name=qname, mode="upsert",
                        source=tbl.get("source"),
                        date_col=tbl.get("date_col"),
                        where_clause=tbl.get("where", ""),
                    )
                    rows = get_row_counts(_db_path, qname)
                    print(f"{qname}: {len(rows)} date buckets loaded")

                    # Also load column metadata if available
                    col_csv = os.path.join(folder, f"{qname}_columns.csv")
                    if os.path.exists(col_csv):
                        columns = {}
                        with open(col_csv, 'r', newline='') as f:
                            reader = csv_mod.DictReader(f)
                            for row in reader:
                                # Strip whitespace from keys to handle \r\n line endings
                                row = {k.strip(): v for k, v in row.items()}
                                col_name = row.get('column_name') or row.get('COLUMN_NAME', '')
                                dtype = row.get('data_type') or row.get('DATA_TYPE', '')
                                if col_name:
                                    columns[col_name] = dtype
                        if columns:
                            insert_column_meta(_db_path, qname, columns, source=tbl.get("source"))
                            print(f"{qname}: {len(columns)} columns loaded")
            else:
                from ..loader import load_precomputed_col_stats
                from ..db import get_metadata
                for tbl in tables:
                    qname = qualified_name(tbl)
                    csv_path = os.path.join(folder, f"{qname}_col.csv")
                    if not os.path.exists(csv_path):
                        continue
                    table_vintage = tbl.get("vintage")
                    if not table_vintage:
                        meta = get_metadata(_db_path, qname)
                        table_vintage = (meta.get("vintage") or "day") if meta else "day"
                    count = load_precomputed_col_stats(
                        db_path=_db_path, csv_path=csv_path,
                        table_name=qname, mode="upsert",
                        source=tbl.get("source"), vintage=table_vintage,
                    )
                    print(f"{qname}: {count} stat rows loaded")

        return {"ok": True, "output": buf.getvalue()}
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": str(e), "output": buf.getvalue()
        })


@app.post("/api/compare/row")
async def api_compare_row(request: Request):
    """Run compare-row and return results."""
    body = await request.json()
    from_date = body.get("from_date")
    to_date = body.get("to_date")

    buf = io.StringIO()
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            from argparse import Namespace
            from ..cli import cmd_compare_row
            args = Namespace(
                project_db=_db_path, config=_config_path,
                from_date=from_date, to_date=to_date,
                yes=True, html=None, title=None, subtitle=None,
            )
            cmd_compare_row(args)
        return {"ok": True, "output": buf.getvalue()}
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": str(e), "output": buf.getvalue()
        })


@app.post("/api/compare/col")
async def api_compare_col(request: Request):
    """Run compare-col and return results."""
    body = await request.json()
    from_date = body.get("from_date")
    to_date = body.get("to_date")

    buf = io.StringIO()
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            from argparse import Namespace
            from ..cli import cmd_compare_col
            args = Namespace(
                project_db=_db_path, config=_config_path,
                from_date=from_date, to_date=to_date,
                no_date_filter=False, vintage=None,
                yes=True, html=None, title=None, subtitle=None,
            )
            cmd_compare_col(args)
        return {"ok": True, "output": buf.getvalue()}
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": str(e), "output": buf.getvalue()
        })


@app.post("/api/query")
async def api_query(request: Request):
    """Run a SQL query against the database (read + write)."""
    body = await request.json()
    sql = body.get("sql", "").strip()

    if not sql:
        return JSONResponse(status_code=400, content={"error": "Empty SQL"})

    try:
        conn = sqlite3.connect(_db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute(sql)

        first_word = sql.split()[0].upper() if sql else ""
        if first_word in ("SELECT", "PRAGMA", "EXPLAIN", "WITH"):
            rows = cursor.fetchall()
            conn.close()
            if not rows:
                return {"columns": [], "rows": []}
            columns = list(rows[0].keys())
            data = [dict(row) for row in rows]
            return {"columns": columns, "rows": data}
        else:
            rowcount = cursor.rowcount
            conn.commit()
            conn.close()
            return {"columns": [], "rows": [],
                    "message": f"{first_word} executed ({rowcount} row(s) affected)"}
    except sqlite3.Error as e:
        return JSONResponse(status_code=400, content={"error": str(e)})


@app.get("/api/report/{report_type}")
async def api_report(report_type: str):
    """Serve generated HTML report."""
    output_dir = _resolve("output")
    filename = f"compare_{report_type}.html"
    path = os.path.join(output_dir, filename)
    if os.path.exists(path):
        return FileResponse(path, media_type="text/html")
    return JSONResponse(status_code=404, content={"error": f"Report not found: {filename}"})


# ---------------------------------------------------------------------------
# Config management
# ---------------------------------------------------------------------------

@app.post("/api/migrate/qnames")
async def api_migrate_qnames():
    """Reconcile _table_pairs / _column_meta / _col_stats / _row_counts /
    _metadata / _sample_date with the per-side qnames derived from the
    current config.

    Prior to the per-side `name` derivation fix, pairs whose left.table !=
    right.table == pair_name collapsed both sides to `{source}_{pair_name}`,
    losing one side's stored rows. This endpoint:
      1. Recomputes the correct (table_left, table_right) for every pair
         from the config using _derive_side_name.
      2. Updates _table_pairs to point at the new qnames.
      3. For each qname table (_column_meta, _col_stats, _row_counts,
         _metadata, _sample_date), renames rows from the old qname to the
         new one when the new qname is otherwise empty (never overwrites).
      4. Returns a per-pair report so the user knows which sides still
         need a fresh extract.
    """
    import sqlite3
    try:
        from ..config import load_unified_config, _derive_side_name
        from ..platforms.base import qualified_name

        config = load_unified_config(_config_path)
        conn = sqlite3.connect(_db_path)
        cursor = conn.cursor()

        # Tables keyed on a single qname column
        qname_tables = [
            ("_column_meta", "source_table"),
            ("_col_stats", "source_table"),
            ("_row_counts", "source_table"),
            ("_metadata", "table_name"),
            ("_sample_date", "table_name"),
        ]

        report = []
        for pair_name, pair_cfg in config.get("pairs", {}).items():
            left = dict(pair_cfg.get("left", {}))
            right = dict(pair_cfg.get("right", {}))
            left["name"] = _derive_side_name(left, pair_name)
            right["name"] = _derive_side_name(right, pair_name)
            new_left = qualified_name(left)
            new_right = qualified_name(right)

            cursor.execute(
                "SELECT table_left, table_right FROM _table_pairs WHERE pair_name=?",
                (pair_name,),
            )
            row = cursor.fetchone()
            if not row:
                report.append({
                    "pair": pair_name,
                    "status": "not_registered",
                    "new_left": new_left, "new_right": new_right,
                })
                continue
            old_left, old_right = row

            pair_report = {
                "pair": pair_name,
                "left":  {"old": old_left,  "new": new_left},
                "right": {"old": old_right, "new": new_right},
                "renamed": [], "skipped": [],
            }

            for side_label, old_q, new_q in (
                ("left",  old_left,  new_left),
                ("right", old_right, new_right),
            ):
                if old_q == new_q:
                    continue
                for tbl, col in qname_tables:
                    cursor.execute(f"SELECT COUNT(*) FROM {tbl} WHERE {col}=?", (old_q,))
                    n_old = cursor.fetchone()[0]
                    if not n_old:
                        continue
                    cursor.execute(f"SELECT COUNT(*) FROM {tbl} WHERE {col}=?", (new_q,))
                    n_new = cursor.fetchone()[0]
                    if n_new:
                        pair_report["skipped"].append({
                            "table": tbl, "side": side_label,
                            "reason": f"{new_q} already has {n_new} rows",
                        })
                        continue
                    cursor.execute(
                        f"UPDATE {tbl} SET {col}=? WHERE {col}=?", (new_q, old_q),
                    )
                    pair_report["renamed"].append({
                        "table": tbl, "side": side_label, "n_rows": n_old,
                    })

            if old_left != new_left or old_right != new_right:
                cursor.execute(
                    "UPDATE _table_pairs SET table_left=?, table_right=? WHERE pair_name=?",
                    (new_left, new_right, pair_name),
                )
            pair_report["status"] = "migrated" if (
                pair_report["renamed"] or old_left != new_left or old_right != new_right
            ) else "already_clean"
            report.append(pair_report)

        conn.commit()
        conn.close()
        return {"ok": True, "pairs": report}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/api/config")
async def api_get_config():
    """Return the current config JSON."""
    try:
        with open(_config_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return JSONResponse(status_code=404, content={"error": "Config not found"})


@app.post("/api/config/upload")
async def api_upload_config(file: UploadFile = File(...)):
    """Upload and replace the config JSON file.

    Also registers pairs and metadata into the database.
    """
    try:
        content = await file.read()
        config = json.loads(content)

        # Validate it has a pairs section
        if "pairs" not in config:
            return JSONResponse(status_code=400, content={
                "error": "Invalid config: missing 'pairs' key"
            })

        # Write to disk
        with open(_config_path, 'w') as f:
            json.dump(config, f, indent=2)

        # Register pairs into the database
        registered = _sync_config_pairs(config)

        return {"ok": True, "pairs": len(config["pairs"]), "registered": registered}
    except json.JSONDecodeError as e:
        return JSONResponse(status_code=400, content={"error": f"Invalid JSON: {e}"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/config")
async def api_update_config(request: Request):
    """Partial update to config (merge into existing)."""
    body = await request.json()
    try:
        from ..config import load_unified_config, save_unified_config
        config = load_unified_config(_config_path)

        # Merge pairs
        if "pairs" in body:
            for pname, pcfg in body["pairs"].items():
                if pname in config["pairs"]:
                    config["pairs"][pname].update(pcfg)
                else:
                    config["pairs"][pname] = pcfg

        # Merge top-level metadata, settings, and date_types
        if "metadata" in body:
            config.setdefault("metadata", {}).update(body["metadata"])
        if "settings" in body:
            config.setdefault("settings", {}).update(body["settings"])
        if "date_types" in body:
            config.setdefault("date_types", {}).update(body["date_types"])

        save_unified_config(config, _config_path)
        _sync_config_pairs(config)

        return {"ok": True}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Pair management (new UI)
# ---------------------------------------------------------------------------

@app.post("/api/pairs/reload")
async def api_reload_pairs():
    """Wipe all pairs and reload from the current config JSON on disk."""
    try:
        from ..config import load_unified_config
        config = load_unified_config(_config_path)

        # Wipe existing pairs in config, then re-read from file
        # (this forces a clean slate from whatever config file is active)
        n_pairs = len(config.get("pairs", {}))

        # Re-sync into database
        registered = _sync_config_pairs(config)

        return {"ok": True, "pairs": n_pairs, "registered": registered,
                "config_path": _config_path}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/pairs/list")
async def api_list_pairs():
    """List all pairs with their full configuration."""
    try:
        from ..config import load_unified_config
        config = load_unified_config(_config_path)

        pairs_list = []
        for pair_name, pair_cfg in config.get("pairs", {}).items():
            # Normalize processed: list → newline-joined string for web UI
            for side in ("left", "right"):
                side_cfg = pair_cfg.get(side, {})
                p = side_cfg.get("processed", "")
                if isinstance(p, list):
                    side_cfg["processed"] = "\n".join(p)
            # col_filter lives in _table_pairs (not config). Fetch per-pair.
            from ..db import get_pair_col_filter
            cf = get_pair_col_filter(_db_path, pair_name)
            pairs_list.append({
                "name": pair_name,
                "description": pair_cfg.get("description", ""),
                "left": pair_cfg.get("left", {}),
                "right": pair_cfg.get("right", {}),
                "mode": pair_cfg.get("mode", "incremental"),
                "vintage": pair_cfg.get("vintage", ""),
                "col_filter": cf if (cf.get("include") or cf.get("exclude")) else None,
                "dateRangeMode": pair_cfg.get("dateRangeMode", "global"),
                "fromDate": pair_cfg.get("fromDate", ""),
                "toDate": pair_cfg.get("toDate", ""),
                "excludeDates": pair_cfg.get("excludeDates", []),
                "overlap": pair_cfg.get("overlap", 7),
                "lastLoaded": pair_cfg.get("lastLoaded", ""),
                "skip": bool(pair_cfg.get("skip", False)),
                "selected": not bool(pair_cfg.get("skip", False)),
                "expanded": False,
                "validated": True,
            })

        return {"pairs": pairs_list}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/pairs/skip")
async def api_set_pair_skip(request: Request):
    """Toggle the persisted `skip` flag for one or more pairs.

    Body shapes:
      {"pair_name": "P1", "skip": true}            -- single pair
      {"pairs": {"P1": false, "P2": true}}         -- bulk
    A pair with skip=true is hidden from /load_row, /row_compare,
    /load_col, /col_compare, /col_mapping. The pairs page is the only
    place to toggle it back on.
    """
    body = await request.json()
    try:
        from ..config import load_unified_config, save_unified_config
        config = load_unified_config(_config_path)
        pairs_cfg = config.setdefault("pairs", {})

        # Normalize input to a {name: bool} dict.
        if "pairs" in body and isinstance(body["pairs"], dict):
            updates = {k: bool(v) for k, v in body["pairs"].items()}
        elif "pair_name" in body and "skip" in body:
            updates = {body["pair_name"]: bool(body["skip"])}
        else:
            return JSONResponse(status_code=400, content={"error": "Expected {pair_name, skip} or {pairs: {...}}"})

        unknown = [n for n in updates if n not in pairs_cfg]
        if unknown:
            return JSONResponse(status_code=404, content={"error": f"Unknown pairs: {unknown}"})

        for name, skip in updates.items():
            if skip:
                pairs_cfg[name]["skip"] = True
            else:
                # Drop the key entirely when re-enabling, to keep dtrack.json clean.
                pairs_cfg[name].pop("skip", None)

        save_unified_config(_config_path, config)
        return {"ok": True, "updated": list(updates.keys())}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/pairs/{pair_name}")
async def api_update_pair(pair_name: str, request: Request):
    """Update an existing pair configuration."""
    body = await request.json()

    try:
        from ..config import load_unified_config, save_unified_config
        from ..platforms.base import qualified_name
        from ..db import register_table_pair, update_metadata
        config = load_unified_config(_config_path)

        if pair_name not in config.get("pairs", {}):
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        pair_cfg = config["pairs"][pair_name]

        # Update only valid pair-config keys (not frontend-only fields like pair_name).
        # col_filter is dual-written: the canonical store is _table_pairs in the DB
        # (see below), but we also persist a copy in dtrack.json so the config file
        # reflects exactly what the UI sees -- otherwise users edit include/exclude
        # in col_gen, save, and wonder why the json doesn't change.
        _PAIR_KEYS = {"left", "right", "col_map", "col_filter", "description", "mode",
                       "vintage",
                       "fromDate", "toDate", "dateRangeMode", "overlap",
                       "skip", "ignore_rows", "ignore_columns",
                       "col_type_overrides", "where_map", "time_map",
                       "comment_map", "diff_map", "excludeDates"}
        for key in _PAIR_KEYS:
            if key in body:
                # Normalize an explicit null col_filter into absence (keeps json clean).
                if key == "col_filter" and body[key] is None:
                    pair_cfg.pop("col_filter", None)
                else:
                    pair_cfg[key] = body[key]

        for side in ("left", "right"):
            if side in pair_cfg:
                _prepare_side(pair_cfg[side])

        save_unified_config(config, _config_path)

        # Inject name for DB registration (auto-derived, not persisted).
        # Each side derives its own name from `table` so left/right qnames
        # don't collide when both sides share a source.
        from ..config import _derive_side_name
        left = pair_cfg.get("left", {})
        right = pair_cfg.get("right", {})
        left["name"] = _derive_side_name(left, pair_name)
        right["name"] = _derive_side_name(right, pair_name)
        if left.get("source") and right.get("source"):
            table_left = qualified_name(left)
            table_right = qualified_name(right)
            register_table_pair(
                _db_path, pair_name, table_left, table_right,
                source_left=left.get("source"),
                source_right=right.get("source"),
                col_mappings=pair_cfg.get("col_map") or None,
            )
            pair_vintage = pair_cfg.get("vintage", "")
            for tbl_cfg in [left, right]:
                qname = qualified_name(tbl_cfg)
                # Use patch to avoid wiping existing fields (min/max dates, etc.)
                from ..db import patch_metadata, get_metadata as _get_meta
                existing = _get_meta(_db_path, qname)
                if existing:
                    patch_fields = {}
                    if tbl_cfg.get("source"):
                        patch_fields["source"] = tbl_cfg["source"]
                    if tbl_cfg.get("table"):
                        patch_fields["source_table"] = tbl_cfg["table"]
                    if tbl_cfg.get("date_col"):
                        patch_fields["date_var"] = tbl_cfg["date_col"]
                    if pair_vintage:
                        patch_fields["vintage"] = pair_vintage
                    if patch_fields:
                        patch_metadata(_db_path, qname, **patch_fields)
                else:
                    update_metadata(_db_path, {
                        "table_name": qname,
                        "source": tbl_cfg.get("source"),
                        "source_table": tbl_cfg.get("table"),
                        "date_var": tbl_cfg.get("date_col"),
                        "vintage": pair_vintage,
                        "data_type": "row",
                    })

        # Save col_filter AFTER register_table_pair — so even if that path
        # somehow wipes the columns again, our write is the last one and
        # persists. Surface errors to the frontend instead of swallowing.
        if "col_filter" in body:
            from ..db import save_pair_col_filter
            cf = body.get("col_filter") or {}
            save_pair_col_filter(
                _db_path, pair_name,
                include=cf.get("include"),
                exclude=cf.get("exclude"),
            )
            print(f"[col_filter] {pair_name}: include={cf.get('include')} exclude={cf.get('exclude')}")

        return {"ok": True, "pair_name": pair_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/scan/csv")
async def api_scan_csv(dir: str = "./output/", type: str = "row"):
    """Scan directory for CSV files."""
    try:
        from pathlib import Path

        dir_path = Path(_resolve(dir))
        if not dir_path.exists():
            return {"files": []}

        pattern = f"*_{type}.csv"
        files = []

        for csv_file in dir_path.glob(pattern):
            stats = csv_file.stat()
            parts = csv_file.stem.split('_')

            # Parse filename: "customer_daily_left_row.csv"
            side = 'left' if 'left' in parts else 'right' if 'right' in parts else 'unknown'
            pair = '_'.join(p for p in parts if p not in ['left', 'right', 'row', 'col'])

            # Count rows and dates (basic check)
            try:
                with open(csv_file) as f:
                    lines = f.readlines()
                    rows = len(lines) - 1  # minus header
                    dates = rows  # Approximate for now
            except:
                rows = 0
                dates = 0

            from datetime import datetime
            modified = datetime.fromtimestamp(stats.st_mtime)
            time_diff = datetime.now() - modified
            if time_diff.seconds < 60:
                modified_str = f"{time_diff.seconds}s ago"
            elif time_diff.seconds < 3600:
                modified_str = f"{time_diff.seconds // 60}m ago"
            elif time_diff.days == 0:
                modified_str = f"{time_diff.seconds // 3600}h ago"
            else:
                modified_str = f"{time_diff.days}d ago"

            files.append({
                "file": csv_file.name,
                "path": str(csv_file),
                "pair": pair,
                "side": side,
                "rows": rows,
                "dates": dates,
                "modified": modified_str,
                "selected": True
            })

        return {"files": files}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/load/row")
async def api_load_row_files(request: Request):
    """Load row CSV files into database."""
    body = await request.json()
    files = body.get("files", [])
    mode = body.get("mode", "upsert")

    try:
        from ..loader import load_row_counts
        from ..config import load_unified_config, get_all_tables_from_unified
        from ..platforms.base import qualified_name

        config = load_unified_config(_config_path)
        loaded = 0

        for file_path in files:
            # Infer table info from filename
            filename = Path(file_path).stem

            # Skip non-row CSV files (columns, col, etc.)
            if not filename.lower().endswith('_row'):
                continue

            # Find matching table in config
            for tbl in get_all_tables_from_unified(config):
                qname = qualified_name(tbl)
                if qname.lower() in filename.lower():
                    load_row_counts(
                        db_path=_db_path,
                        file_or_folder=file_path,
                        table_name=qname,
                        mode=mode,
                        source=tbl.get("source"),
                        date_col=tbl.get("date_col"),
                        where_clause=tbl.get("where", ""),
                    )
                    loaded += 1
                    break

        return {"ok": True, "loaded": loaded}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/load/row/upload")
async def api_load_row_upload(
    file: UploadFile = File(...),
    table_name: str = Form(""),
    mode: str = Form("upsert"),
):
    """Upload a CSV file and load row counts into a specific table.

    Accepts multipart form data with:
      - file: the CSV file
      - table_name: target table name (e.g. 'oracle_cust_daily')
      - mode: 'upsert' (default) or 'replace'
    """
    import tempfile

    if not table_name:
        return JSONResponse(status_code=400, content={"error": "table_name is required"})

    try:
        # Write uploaded file to temp location
        content = await file.read()
        with tempfile.NamedTemporaryFile(
            mode='wb', suffix='.csv', delete=False
        ) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        # Count existing rows for this table to compute new vs updated
        from ..db import get_row_counts, get_metadata
        existing_rows = {}
        try:
            for dt, count in get_row_counts(_db_path, table_name):
                existing_rows[dt] = count
        except Exception:
            pass  # table may not exist yet

        # Find table config for metadata
        from ..config import load_unified_config, get_all_tables_from_unified
        from ..platforms.base import qualified_name

        config = load_unified_config(_config_path)
        tbl_cfg = None
        for tbl in get_all_tables_from_unified(config):
            if qualified_name(tbl).lower() == table_name.lower():
                tbl_cfg = tbl
                break

        from ..loader import load_row_counts
        # date_col=None lets the CSV parser auto-detect (date_value, dt, date, etc.)
        # date_var_override stores the source DB column name in metadata
        load_row_counts(
            db_path=_db_path,
            file_or_folder=tmp_path,
            table_name=table_name,
            mode=mode,
            source=tbl_cfg.get("source") if tbl_cfg else None,
            date_col=None,
            date_var_override=tbl_cfg.get("date_col") if tbl_cfg else None,
        )

        # Compute new vs updated dates
        new_rows = {}
        try:
            for dt, count in get_row_counts(_db_path, table_name):
                new_rows[dt] = count
        except Exception:
            pass

        new_dates = len(set(new_rows.keys()) - set(existing_rows.keys()))
        updated_dates = len(set(new_rows.keys()) & set(existing_rows.keys()))

        # Clean up temp file
        os.unlink(tmp_path)

        return {
            "ok": True,
            "loaded": len(new_rows),
            "new_dates": new_dates,
            "updated_dates": updated_dates,
        }
    except Exception as e:
        import traceback
        return JSONResponse(status_code=500, content={
            "ok": False,
            "error": f"{type(e).__name__}: {e}",
        })


@app.post("/api/load/columns/upload")
async def api_load_columns_upload(
    file: UploadFile = File(...),
    table_name: str = Form(""),
    source: str = Form(""),
):
    """Upload a column metadata CSV and load into _column_meta.

    CSV should have headers: COLUMN_NAME, DATA_TYPE (or column_name, data_type).
    """
    import csv as csv_mod

    if not table_name:
        return JSONResponse(status_code=400, content={"error": "table_name is required"})

    try:
        from ..db import insert_column_meta

        content = await file.read()
        text = content.decode("utf-8")
        reader = csv_mod.DictReader(io.StringIO(text))
        columns = {}
        for row in reader:
            # Strip whitespace from keys to handle \r\n line endings
            row = {k.strip(): v for k, v in row.items()}
            col_name = row.get('column_name') or row.get('COLUMN_NAME', '')
            dtype = row.get('data_type') or row.get('DATA_TYPE', '')
            if col_name:
                columns[col_name] = dtype

        if not columns:
            return JSONResponse(status_code=400, content={"error": "No columns found in CSV"})

        count = insert_column_meta(_db_path, table_name, columns, source=source or None)
        return {"ok": True, "loaded": count, "table_name": table_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/load/columns/path")
async def api_load_columns_path(request: Request):
    """Load a server-side column metadata CSV by path.

    Body: {path, table_name, source}
    """
    import csv as csv_mod

    body = await request.json()
    csv_path = body.get("path", "")
    if csv_path and not os.path.isabs(csv_path):
        csv_path = _resolve(csv_path)
    table_name = body.get("table_name", "")
    source = body.get("source", "")

    if not table_name:
        return JSONResponse(status_code=400, content={"error": "table_name is required"})
    if not csv_path or not os.path.exists(csv_path):
        return JSONResponse(status_code=400, content={"error": f"File not found: {csv_path}"})

    try:
        from ..db import insert_column_meta

        columns = {}
        with open(csv_path, 'r', newline='') as f:
            reader = csv_mod.DictReader(f)
            for row in reader:
                row = {k.strip(): v for k, v in row.items()}
                col_name = row.get('column_name') or row.get('COLUMN_NAME', '')
                dtype = row.get('data_type') or row.get('DATA_TYPE', '')
                if col_name:
                    columns[col_name] = dtype

        if not columns:
            return JSONResponse(status_code=400, content={"error": "No columns found in CSV"})

        count = insert_column_meta(_db_path, table_name, columns, source=source or None)
        return {"ok": True, "loaded": count, "table_name": table_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/load/row/path")
async def api_load_row_path(request: Request):
    """Load a server-side CSV file by path into a specific table.

    Body: {path, table_name, mode}
    """
    body = await request.json()
    csv_path = body.get("path", "")
    if csv_path and not os.path.isabs(csv_path):
        csv_path = _resolve(csv_path)
    table_name = body.get("table_name", "")
    mode = body.get("mode", "upsert")

    if not csv_path or not table_name:
        return JSONResponse(status_code=400, content={
            "error": "path and table_name are required"
        })

    if not os.path.exists(csv_path):
        return JSONResponse(status_code=404, content={
            "error": f"File not found: {csv_path}"
        })

    try:
        from ..db import get_row_counts
        from ..config import load_unified_config, get_all_tables_from_unified
        from ..platforms.base import qualified_name

        # Count existing rows
        existing_rows = {}
        try:
            for dt, count in get_row_counts(_db_path, table_name):
                existing_rows[dt] = count
        except Exception:
            pass

        # Find table config
        config = load_unified_config(_config_path)
        tbl_cfg = None
        for tbl in get_all_tables_from_unified(config):
            if qualified_name(tbl).lower() == table_name.lower():
                tbl_cfg = tbl
                break

        from ..loader import load_row_counts
        # date_col=None lets the CSV parser auto-detect (date_value, dt, date, etc.)
        load_row_counts(
            db_path=_db_path,
            file_or_folder=csv_path,
            table_name=table_name,
            mode=mode,
            source=tbl_cfg.get("source") if tbl_cfg else None,
            date_col=None,
            date_var_override=tbl_cfg.get("date_col") if tbl_cfg else None,
        )

        # Compute new vs updated
        new_rows = {}
        try:
            for dt, count in get_row_counts(_db_path, table_name):
                new_rows[dt] = count
        except Exception:
            pass

        new_dates = len(set(new_rows.keys()) - set(existing_rows.keys()))
        updated_dates = len(set(new_rows.keys()) & set(existing_rows.keys()))

        return {
            "ok": True,
            "loaded": len(new_rows),
            "new_dates": new_dates,
            "updated_dates": updated_dates,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/api/scan/folder")
async def api_scan_folder(dir: str = "./csv/"):
    """Scan a server-side folder for CSV files and return file info."""
    try:
        dir_path = Path(_resolve(dir))
        if not dir_path.exists():
            return JSONResponse(status_code=404, content={
                "error": f"Folder not found: {dir}"
            })

        files = []
        for csv_file in sorted(dir_path.glob("*.csv")):
            # Read first few lines for row count and date range
            rows = 0
            min_date = ''
            max_date = ''
            try:
                with open(csv_file, 'r') as f:
                    import csv as csv_mod
                    reader = csv_mod.DictReader(f)
                    headers = [h.lower().strip() for h in (reader.fieldnames or [])]
                    date_aliases = {'dt', 'date', 'rpg_dt', 'eff_dt', 'run_date',
                                    'snap_dt', 'snapshot_dt', 'date_value'}
                    date_col = None
                    for h in headers:
                        if h in date_aliases:
                            date_col = h
                            break

                    dates = []
                    for row in reader:
                        rows += 1
                        if date_col and row.get(date_col):
                            dates.append(row[date_col].strip())

                    if dates:
                        dates.sort()
                        min_date = dates[0]
                        max_date = dates[-1]
            except Exception:
                pass

            files.append({
                "name": csv_file.name,
                "path": str(csv_file.resolve()),
                "rows": rows,
                "minDate": min_date,
                "maxDate": max_date,
            })

        return {"files": files}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Col Stats Load API
# ---------------------------------------------------------------------------

@app.post("/api/load/col-stats/upload")
async def api_load_col_stats_upload(
    file: UploadFile = File(...),
    table_name: str = Form(""),
    mode: str = Form("upsert"),
):
    """Upload pre-computed column stats CSV.

    CSV should have: column_name, dt, col_type, n_total, n_missing, n_unique,
    mean, std, min_val, max_val, top_10
    """
    import tempfile

    if not table_name:
        return JSONResponse(status_code=400, content={"error": "table_name is required"})

    try:
        content = await file.read()
        with tempfile.NamedTemporaryFile(
            mode='wb', suffix='.csv', delete=False
        ) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        from ..loader import load_precomputed_col_stats
        from ..config import load_unified_config, get_all_tables_from_unified
        from ..platforms.base import qualified_name

        # Find table config for vintage info
        config = load_unified_config(_config_path)
        tbl_cfg = None
        for tbl in get_all_tables_from_unified(config):
            if qualified_name(tbl).lower() == table_name.lower():
                tbl_cfg = tbl
                break

        table_vintage = "day"
        if tbl_cfg:
            table_vintage = tbl_cfg.get("vintage") or "day"
        else:
            from ..db import get_metadata
            meta = get_metadata(_db_path, table_name)
            if meta:
                table_vintage = meta.get("vintage") or "day"

        count = load_precomputed_col_stats(
            db_path=_db_path,
            csv_path=tmp_path,
            table_name=table_name,
            mode=mode,
            source=tbl_cfg.get("source") if tbl_cfg else None,
            vintage=table_vintage,
        )

        os.unlink(tmp_path)

        return {"ok": True, "loaded": count, "table_name": table_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": f"{type(e).__name__}: {e}",
        })


@app.post("/api/load/col-stats/path")
async def api_load_col_stats_path(request: Request):
    """Load col stats from server-side path.

    Body: {path, table_name, mode}
    """
    body = await request.json()
    csv_path = body.get("path", "")
    if csv_path and not os.path.isabs(csv_path):
        csv_path = _resolve(csv_path)
    table_name = body.get("table_name", "")
    mode = body.get("mode", "upsert")

    if not table_name:
        return JSONResponse(status_code=400, content={"error": "table_name is required"})
    if not csv_path or not os.path.exists(csv_path):
        return JSONResponse(status_code=400, content={"error": f"File not found: {csv_path}"})

    try:
        from ..loader import load_precomputed_col_stats
        from ..config import load_unified_config, get_all_tables_from_unified
        from ..platforms.base import qualified_name
        from ..db import get_metadata

        config = load_unified_config(_config_path)
        tbl_cfg = None
        for tbl in get_all_tables_from_unified(config):
            if qualified_name(tbl).lower() == table_name.lower():
                tbl_cfg = tbl
                break

        table_vintage = "day"
        if tbl_cfg:
            table_vintage = tbl_cfg.get("vintage") or "day"
        else:
            meta = get_metadata(_db_path, table_name)
            if meta:
                table_vintage = meta.get("vintage") or "day"

        count = load_precomputed_col_stats(
            db_path=_db_path,
            csv_path=csv_path,
            table_name=table_name,
            mode=mode,
            source=tbl_cfg.get("source") if tbl_cfg else None,
            vintage=table_vintage,
        )

        return {"ok": True, "loaded": count, "table_name": table_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/api/status/col")
async def api_status_col():
    """Return col stats summary per table (count, date range, distinct dates)."""
    from ..date_utils import parse_date
    try:
        conn = sqlite3.connect(_db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        status = {}
        try:
            # Get count per table
            cursor.execute(
                "SELECT source_table, COUNT(*) as cnt FROM _col_stats GROUP BY source_table"
            )
            for row in cursor.fetchall():
                status[row["source_table"]] = {"count": row["cnt"]}

            # Get all distinct dates per table for full history
            cursor.execute(
                "SELECT source_table, dt FROM _col_stats GROUP BY source_table, dt"
            )
            dates_by_table = {}
            for row in cursor.fetchall():
                tbl = row["source_table"]
                raw_dt = row["dt"]
                if tbl not in dates_by_table:
                    dates_by_table[tbl] = []
                try:
                    dates_by_table[tbl].append(parse_date(raw_dt))
                except ValueError:
                    dates_by_table[tbl].append(raw_dt)

            for tbl, dates in dates_by_table.items():
                dates.sort()
                if tbl in status:
                    status[tbl]["dates"] = dates
                    status[tbl]["min_date"] = dates[0] if dates else None
                    status[tbl]["max_date"] = dates[-1] if dates else None

        except sqlite3.OperationalError:
            pass  # _col_stats may not exist

        conn.close()
        return {"status": status}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Col Compare API
# ---------------------------------------------------------------------------

@app.put("/api/compare/col/{pair_name}")
async def api_save_col_comparison(pair_name: str, request: Request):
    """Save col comparison state to _col_comparison."""
    body = await request.json()
    try:
        from ..db import save_col_comparison
        save_col_comparison(
            _db_path, pair_name,
            columns_compared=body.get("columns_compared", []),
            matched_columns=body.get("matched_columns", []),
            diff_columns=body.get("diff_columns", []),
        )
        return {"ok": True}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/compare/col/{pair_name}")
async def api_compare_col_pair(pair_name: str, from_date: str = "", to_date: str = ""):
    """Compare col stats for a single pair — returns per-vintage comparisons."""
    try:
        from ..db import (
            get_table_pair, get_col_stats, get_pair_col_map_from_db,
            get_column_meta, get_metadata,
        )
        from ..config import load_unified_config, get_col_type_overrides
        from ..compare import compare_column_stats, resolve_col_type, _has_col_differences
        from ..date_utils import parse_date

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        table_left = pair["table_left"]
        table_right = pair["table_right"]

        config = load_unified_config(_config_path)
        col_type_overrides = get_col_type_overrides(config, pair_name)

        # Col mappings
        col_mappings = get_pair_col_map_from_db(_db_path, pair_name)
        if not col_mappings:
            left_meta = get_column_meta(_db_path, table_left)
            right_meta = get_column_meta(_db_path, table_right)
            left_cols = {c["column_name"]: c.get("data_type", "") for c in left_meta}
            right_cols = {c["column_name"]: c.get("data_type", "") for c in right_meta}
            from ..compare import match_columns_from_dicts
            auto = match_columns_from_dicts(left_cols, right_cols)
            col_mappings = auto.get("matched", {})

        # Full comparison (all dates, like CLI compare-col)
        comparison = compare_column_stats(
            _db_path, table_left, table_right,
            columns=list(col_mappings.keys()) if col_mappings else None,
            col_mappings=col_mappings,
            from_date=from_date or None,
            to_date=to_date or None,
            col_type_overrides=col_type_overrides,
        )

        # Get all left/right column names for left-only/right-only
        stats_left = get_col_stats(_db_path, table_left,
                                   from_date=from_date or None, to_date=to_date or None)
        stats_right = get_col_stats(_db_path, table_right,
                                    from_date=from_date or None, to_date=to_date or None)
        all_left_cols = set(s["column_name"] for s in stats_left)
        all_right_cols = set(s["column_name"] for s in stats_right)
        mapped_left = set(col_mappings.keys())
        mapped_right = set(col_mappings.values())

        # Organize by vintage: {vintage_label: {col_name: comparison_entry}}
        vintages = {}  # ordered list of {label, dt, columns: [...]}
        vintage_order = []  # maintain order by dt

        # Collect all vintages and per-vintage column comparisons
        n_diff_cols = 0
        n_type_mismatch = 0
        diff_col_names = set()

        for left_col, comps in comparison.items():
            for comp in comps:
                vl = comp.get("vintage_label") or comp.get("dt", "")
                dt = comp.get("dt", "")
                if vl not in vintages:
                    vintages[vl] = {"label": vl, "dt": dt, "columns": []}
                    vintage_order.append(vl)

                has_diff = _has_col_differences(comp)
                if has_diff:
                    diff_col_names.add(left_col)

                if comp.get("col_type_left") and comp.get("col_type_right") and \
                   comp["col_type_left"] != comp["col_type_right"]:
                    n_type_mismatch += 1

                vintages[vl]["columns"].append({
                    "left_col": comp.get("left_col", left_col),
                    "right_col": comp.get("right_col", left_col),
                    "col_type": comp.get("col_type", ""),
                    "col_type_left": comp.get("col_type_left", ""),
                    "col_type_right": comp.get("col_type_right", ""),
                    "type_mismatch": comp.get("col_type_left", "") != comp.get("col_type_right", ""),
                    "has_override": left_col in col_type_overrides,
                    "n_total_left": comp.get("n_total_left", 0),
                    "n_total_right": comp.get("n_total_right", 0),
                    "n_total_diff": comp.get("n_total_diff", 0),
                    "n_missing_left": comp.get("n_missing_left", 0),
                    "n_missing_right": comp.get("n_missing_right", 0),
                    "n_missing_diff": comp.get("n_missing_diff", 0),
                    "n_unique_left": comp.get("n_unique_left", 0),
                    "n_unique_right": comp.get("n_unique_right", 0),
                    "n_unique_diff": comp.get("n_unique_diff", 0),
                    "mean_left": comp.get("mean_left"),
                    "mean_right": comp.get("mean_right"),
                    "mean_match": comp.get("mean_match"),
                    "std_left": comp.get("std_left"),
                    "std_right": comp.get("std_right"),
                    "std_match": comp.get("std_match"),
                    "min_left": comp.get("min_left", ""),
                    "min_right": comp.get("min_right", ""),
                    "max_left": comp.get("max_left", ""),
                    "max_right": comp.get("max_right", ""),
                    "top_10_left": comp.get("top_10_left", ""),
                    "top_10_right": comp.get("top_10_right", ""),
                    "has_diff": has_diff,
                })

        n_diff_cols = len(diff_col_names)

        # Build vintage list in date order
        vintage_list = [vintages[vl] for vl in vintage_order]

        # Left-only and right-only columns (from latest stats)
        def _latest(stats_list):
            by_col = {}
            for s in stats_list:
                col = s["column_name"]
                if col not in by_col or (s.get("dt", "") > by_col[col].get("dt", "")):
                    by_col[col] = s
            return by_col

        left_agg = _latest(stats_left)
        right_agg = _latest(stats_right)

        left_only = [
            {"column_name": c, "col_type": left_agg[c].get("col_type", ""),
             "n_total": left_agg[c].get("n_total", ""), "n_missing": left_agg[c].get("n_missing", ""),
             "n_unique": left_agg[c].get("n_unique", "")}
            for c in sorted(all_left_cols - mapped_left) if c in left_agg
        ]
        right_only = [
            {"column_name": c, "col_type": right_agg[c].get("col_type", ""),
             "n_total": right_agg[c].get("n_total", ""), "n_missing": right_agg[c].get("n_missing", ""),
             "n_unique": right_agg[c].get("n_unique", "")}
            for c in sorted(all_right_cols - mapped_right) if c in right_agg
        ]

        # Metadata
        meta_left = get_metadata(_db_path, table_left) or {}
        meta_right = get_metadata(_db_path, table_right) or {}

        # Annotations (col scope) — surfaced for the per-pair edit UI on
        # /col_compare. Persisted via PUT /api/pairs/{name}/annotations.
        pair_cfg = config.get("pairs", {}).get(pair_name, {})
        col_comment_map = pair_cfg.get("comment_map", {}).get("col", {}) or {}
        col_time_map = pair_cfg.get("time_map", {}).get("col", {}) or {}

        return {
            "pair_name": pair_name,
            "table_left": table_left,
            "table_right": table_right,
            "source_left": pair.get("source_left", ""),
            "source_right": pair.get("source_right", ""),
            "col_type_overrides": col_type_overrides,
            "summary": {
                "n_matched": len(col_mappings),
                "n_diff": n_diff_cols,
                "n_left_only": len(left_only),
                "n_right_only": len(right_only),
                "n_type_mismatch": n_type_mismatch,
                "n_vintages": len(vintage_list),
            },
            "vintages": vintage_list,
            "left_only": left_only,
            "right_only": right_only,
            "meta_left": {
                "vintage": meta_left.get("vintage"),
                "date_var": meta_left.get("date_var"),
            },
            "meta_right": {
                "vintage": meta_right.get("vintage"),
                "date_var": meta_right.get("date_var"),
            },
            "annotations": {
                "comment_left": col_comment_map.get("left", ""),
                "comment_right": col_comment_map.get("right", ""),
                "time_left": col_time_map.get("left", ""),
                "time_right": col_time_map.get("right", ""),
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/compare/col/export/html")
async def api_compare_col_export_html(request: Request):
    """Generate full HTML report for col comparison (like CLI compare-col)."""
    body = await request.json()
    from_date = body.get("from_date", "")
    to_date = body.get("to_date", "")
    title = body.get("title", "")
    subtitle = body.get("subtitle", "")
    # Live UI state per pair: {pair_name: {comment_left, comment_right,
    # time_left, time_right}}. Falls back to pair_cfg when absent.
    live_pairs = body.get("pairs", {}) or {}

    try:
        from ..db import list_table_pairs, get_metadata, get_row_comparison
        from ..config import load_unified_config, get_col_type_overrides
        from ..compare import compare_column_stats
        from ..html_export import generate_column_stats_html, create_column_stats_table, wrap_html_document

        config = load_unified_config(_config_path)
        pairs = list_table_pairs(_db_path)

        row_sections = []
        for pair in pairs:
            pname = pair["pair_name"]
            pair_cfg = config.get("pairs", {}).get(pname, {})
            if pair_cfg.get("skip"):
                continue

            col_type_overrides = get_col_type_overrides(config, pname)
            from ..db import get_pair_col_map_from_db, get_column_meta
            col_mappings = get_pair_col_map_from_db(_db_path, pname)
            if not col_mappings:
                left_meta = get_column_meta(_db_path, pair["table_left"])
                right_meta = get_column_meta(_db_path, pair["table_right"])
                left_cols = {c["column_name"]: c.get("data_type", "") for c in left_meta}
                right_cols = {c["column_name"]: c.get("data_type", "") for c in right_meta}
                from ..compare import match_columns_from_dicts
                auto = match_columns_from_dicts(left_cols, right_cols)
                col_mappings = auto.get("matched", {})

            # No matched_dates filter -- exports show every (col, dt) the col
            # stats have, just like the col_compare page itself. Filtering by
            # row-compare's matching_dates would silently empty the export
            # when the date formats don't line up between row counts and col
            # stats (vintage-bucketed col stats vs raw row dates, etc.).
            comparison = compare_column_stats(
                _db_path, pair["table_left"], pair["table_right"],
                columns=list(col_mappings.keys()) if col_mappings else None,
                col_mappings=col_mappings,
                from_date=from_date or None,
                to_date=to_date or None,
                col_type_overrides=col_type_overrides,
            )

            meta_left = get_metadata(_db_path, pair["table_left"]) or {}
            meta_right = get_metadata(_db_path, pair["table_right"]) or {}
            vintage = meta_left.get("vintage") or meta_right.get("vintage") or "day"

            # Live UI state wins over saved config so edits show up in the
            # downloaded HTML even before the user blurs the input. The
            # frontend posts {pairs: {<pname>: {time_left, time_right,
            # comment_left, comment_right}}} alongside from_date/to_date.
            live = live_pairs.get(pname) or {}
            saved_comment = pair_cfg.get("comment_map", {}).get("col", {}) or {}
            saved_time = pair_cfg.get("time_map", {}).get("col", {}) or {}
            comment_map = {
                "left":  live.get("comment_left",  saved_comment.get("left", "")),
                "right": live.get("comment_right", saved_comment.get("right", "")),
            }
            time_map = {
                "left":  live.get("time_left",  saved_time.get("left", "")),
                "right": live.get("time_right", saved_time.get("right", "")),
            } if (live.get("time_left") or live.get("time_right")
                  or saved_time.get("left") or saved_time.get("right")) else None

            rows_html = generate_column_stats_html(
                pair_name=pname,
                source_left=pair.get("source_left", ""),
                source_right=pair.get("source_right", ""),
                table_left=pair["table_left"],
                table_right=pair["table_right"],
                comparison=comparison,
                col_mappings=col_mappings,
                metadata_left=meta_left,
                metadata_right=meta_right,
                time_map=time_map,
                comment_left=comment_map.get("left", ""),
                comment_right=comment_map.get("right", ""),
            )
            row_sections.append(rows_html)

        table_html = create_column_stats_table(
            row_sections,
            vintage=meta_left.get("vintage") or "day" if pairs else "day",
        )
        doc = wrap_html_document(
            title=title or "Column Statistics Comparison",
            sections=[table_html],
            subtitle=subtitle,
        )

        return HTMLResponse(content=doc)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/compare/col/export/log")
async def api_compare_col_export_log(request: Request):
    """Generate compact comparison log and save to {outdir}/{pair}_col.log files.

    Returns the combined log text.
    """
    body = await request.json()
    from_date = body.get("from_date", "")
    to_date = body.get("to_date", "")

    try:
        from ..db import list_table_pairs, get_metadata, get_row_comparison
        from ..config import load_unified_config, get_col_type_overrides
        from ..compare import compare_column_stats, _has_col_differences
        from ..db import get_pair_col_map_from_db, get_column_meta

        config = load_unified_config(_config_path)
        pairs = list_table_pairs(_db_path)
        outdir = _resolve(config.get("settings", {}).get("aws_outdir", "./csv/"))
        os.makedirs(outdir, exist_ok=True)

        all_logs = []

        for pair in pairs:
            pname = pair["pair_name"]
            pair_cfg = config.get("pairs", {}).get(pname, {})
            if pair_cfg.get("skip"):
                continue

            col_type_overrides = get_col_type_overrides(config, pname)
            col_mappings = get_pair_col_map_from_db(_db_path, pname)
            if not col_mappings:
                left_meta = get_column_meta(_db_path, pair["table_left"])
                right_meta = get_column_meta(_db_path, pair["table_right"])
                left_cols = {c["column_name"]: c.get("data_type", "") for c in left_meta}
                right_cols = {c["column_name"]: c.get("data_type", "") for c in right_meta}
                from ..compare import match_columns_from_dicts
                auto = match_columns_from_dicts(left_cols, right_cols)
                col_mappings = auto.get("matched", {})

            comparison = compare_column_stats(
                _db_path, pair["table_left"], pair["table_right"],
                columns=list(col_mappings.keys()) if col_mappings else None,
                col_mappings=col_mappings,
                from_date=from_date or None,
                to_date=to_date or None,
                col_type_overrides=col_type_overrides,
            )

            meta_left = get_metadata(_db_path, pair["table_left"]) or {}
            meta_right = get_metadata(_db_path, pair["table_right"]) or {}
            vintage = meta_left.get("vintage") or meta_right.get("vintage") or "day"

            log = _format_col_comparison_log(
                pname, pair, comparison, vintage, col_mappings)

            # Write per-pair log file
            log_path = os.path.join(outdir, f"{pname}_col.log")
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(log)

            all_logs.append(log)

        combined = "\n".join(all_logs)
        return {"ok": True, "log": combined, "outdir": outdir}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


def _format_col_comparison_log(pair_name, pair, comparison, vintage, col_mappings):
    """Format compact comparison log with diffs and top_10.

    First vintage emits a top_10 dump for every column (handy for spot-
    checking categorical distributions); subsequent vintages only emit
    top_10 inside DIFF blocks.
    """
    from ..compare import _has_col_differences
    from ..constants import STAT_ROUND_DECIMALS

    def _fmt_round(v):
        if v is None or v == "":
            return "?"
        try:
            return f"{round(float(v), STAT_ROUND_DECIMALS)}"
        except (ValueError, TypeError):
            return str(v)

    def _fmt_top10(t):
        if not t:
            return ""
        s = str(t)
        # JSON-style top_10 from compute_categorical_stats — unwrap to the
        # human-readable "value(count); ..." shape used elsewhere.
        if s.startswith('['):
            try:
                import json as _json
                entries = _json.loads(s)
                return "; ".join(f"{e.get('value','')}({e.get('count','')})" for e in entries)
            except (ValueError, TypeError):
                return s
        return s

    lines = []
    lines.append(f"=== {pair_name} ({pair['table_left']} vs {pair['table_right']}) ===")

    # Organize by vintage
    vintages = {}  # vintage_label -> {col: comp}
    for col_name, comps in comparison.items():
        for comp in comps:
            vl = comp.get("vintage_label") or comp.get("dt", "")
            vintages.setdefault(vl, {})[col_name] = comp

    n_matched = len(col_mappings)
    n_diff_cols = len(set(
        col for col, comps in comparison.items()
        for c in comps if _has_col_differences(c)
    ))
    lines.append(f"  vintage: {vintage} | {len(vintages)} vintages | "
                 f"{n_matched} matched cols, {n_diff_cols} diff")
    lines.append("")

    sorted_vls = sorted(vintages.keys())
    first_vl = sorted_vls[0] if sorted_vls else None

    for vl in sorted_vls:
        col_comps = vintages[vl]
        diffs = {col: c for col, c in col_comps.items() if _has_col_differences(c)}
        is_first = (vl == first_vl)

        if not diffs and not is_first:
            lines.append(f"  MATCH: {vl} — all {len(col_comps)} columns match")
            continue

        if diffs:
            lines.append(f"  DIFF: {vl}")
        else:
            lines.append(f"  MATCH: {vl} — all {len(col_comps)} columns match")

        # Diff details
        for col, c in sorted(diffs.items()):
            parts = []
            for stat in ("n_total", "n_missing", "n_unique"):
                d = c.get(f"{stat}_diff", 0)
                if d != 0:
                    l = c.get(f"{stat}_left", 0)
                    r = c.get(f"{stat}_right", 0)
                    sign = "+" if d > 0 else ""
                    parts.append(f"{stat} {sign}{d} ({l}->{r})")

            if c.get("mean_match") is False:
                parts.append(f"mean ({_fmt_round(c.get('mean_left'))}->{_fmt_round(c.get('mean_right'))})")
            if c.get("std_match") is False:
                parts.append(f"std ({_fmt_round(c.get('std_left'))}->{_fmt_round(c.get('std_right'))})")
            if c.get("min_match") is False:
                parts.append(f"min ({c.get('min_left','?')}->{c.get('min_right','?')})")
            if c.get("max_match") is False:
                parts.append(f"max ({c.get('max_left','?')}->{c.get('max_right','?')})")

            detail = "  ".join(parts) if parts else "values differ"
            lines.append(f"    {col:<20s} {detail}")

            # Diff-only top_10 dump
            t10_l = _fmt_top10(c.get("top_10_left", ""))
            t10_r = _fmt_top10(c.get("top_10_right", ""))
            if t10_l or t10_r:
                if t10_l != t10_r:
                    lines.append(f"      top_10 L: {t10_l[:200]}")
                    lines.append(f"      top_10 R: {t10_r[:200]}")
                else:
                    lines.append(f"      top_10:   {t10_l[:200]}")

        # First vintage: emit top_10 for EVERY column (matched and diff),
        # so the user can eyeball the actual category distributions once.
        if is_first:
            lines.append(f"    --- top_10 dump ({vl}) ---")
            for col in sorted(col_comps.keys()):
                c = col_comps[col]
                t10_l = _fmt_top10(c.get("top_10_left", ""))
                t10_r = _fmt_top10(c.get("top_10_right", ""))
                if not t10_l and not t10_r:
                    continue
                if t10_l == t10_r:
                    lines.append(f"    {col:<20s} top_10:   {t10_l[:200]}")
                else:
                    lines.append(f"    {col:<20s} top_10 L: {t10_l[:200]}")
                    lines.append(f"    {'':<20s} top_10 R: {t10_r[:200]}")

        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Col Compare Excel Export
# ---------------------------------------------------------------------------

@app.get("/api/compare/col/export/excel")
async def api_compare_col_export_excel(from_date: str = "", to_date: str = ""):
    """Generate Excel workbook for col comparison — one sheet per pair, transposed layout."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from ..db import (
            list_table_pairs, get_metadata, get_row_comparison,
            get_pair_col_map_from_db, get_column_meta,
        )
        from ..config import load_unified_config, get_col_type_overrides
        from ..compare import compare_column_stats, _has_col_differences

        config = load_unified_config(_config_path)
        pairs = list_table_pairs(_db_path)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        bold = Font(bold=True)

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        stat_rows = ["Type", "N_Total", "N_Missing", "N_Unique", "Mean", "Std", "Min", "Max"]
        stat_keys_l = ["col_type", "n_total_left", "n_missing_left", "n_unique_left",
                        "mean_left", "std_left", "min_left", "max_left"]
        stat_keys_r = ["col_type", "n_total_right", "n_missing_right", "n_unique_right",
                        "mean_right", "std_right", "min_right", "max_right"]

        for pair in pairs:
            pname = pair["pair_name"]
            pair_cfg = config.get("pairs", {}).get(pname, {})
            if pair_cfg.get("skip"):
                continue

            col_type_overrides = get_col_type_overrides(config, pname)
            col_mappings = get_pair_col_map_from_db(_db_path, pname)
            if not col_mappings:
                left_meta = get_column_meta(_db_path, pair["table_left"])
                right_meta = get_column_meta(_db_path, pair["table_right"])
                left_cols = {c["column_name"]: c.get("data_type", "") for c in left_meta}
                right_cols = {c["column_name"]: c.get("data_type", "") for c in right_meta}
                from ..compare import match_columns_from_dicts
                auto = match_columns_from_dicts(left_cols, right_cols)
                col_mappings = auto.get("matched", {})

            comparison = compare_column_stats(
                _db_path, pair["table_left"], pair["table_right"],
                columns=list(col_mappings.keys()) if col_mappings else None,
                col_mappings=col_mappings,
                from_date=from_date or None, to_date=to_date or None,
                col_type_overrides=col_type_overrides,
            )

            meta_left = get_metadata(_db_path, pair["table_left"]) or {}
            meta_right = get_metadata(_db_path, pair["table_right"]) or {}
            source_left = (pair.get("source_left") or "LEFT").upper()
            source_right = (pair.get("source_right") or "RIGHT").upper()
            table_left = meta_left.get("source_table") or pair["table_left"]
            table_right = meta_right.get("source_table") or pair["table_right"]

            # Organize by vintage
            vintages = {}
            for col_name, comps in comparison.items():
                for comp in comps:
                    vl = comp.get("vintage_label") or comp.get("dt", "")
                    vintages.setdefault(vl, []).append(comp)

            # Sheet name: truncate to 31 chars (Excel limit)
            ws = wb.create_sheet(title=pname[:31])
            row_num = 1

            for vl in sorted(vintages.keys()):
                comps = vintages[vl]
                has_diff = any(_has_col_differences(c) for c in comps)

                # Sort: diff columns first, then matching, alphabetical within each group
                comps.sort(key=lambda c: (0 if _has_col_differences(c) else 1, c.get("left_col", "")))
                col_names_left = [c.get("left_col", "") for c in comps]
                col_names_right = [c.get("right_col", "") or c.get("left_col", "") for c in comps]

                # Track which (col_index, stat_index) have diffs for red highlighting
                diff_cells = set()  # (col_idx, stat_idx)
                problem_cols = set()  # col indices with at least one diff
                for ci, comp in enumerate(comps):
                    if comp.get("n_total_diff", 0) != 0: diff_cells.add((ci, 1))
                    if comp.get("n_missing_diff", 0) != 0: diff_cells.add((ci, 2))
                    if comp.get("n_unique_diff", 0) != 0: diff_cells.add((ci, 3))
                    if comp.get("mean_match") is False: diff_cells.add((ci, 4))
                    if comp.get("std_match") is False: diff_cells.add((ci, 5))
                    if comp.get("min_match") is False: diff_cells.add((ci, 6))
                    if comp.get("max_match") is False: diff_cells.add((ci, 7))
                for ci, si in diff_cells:
                    problem_cols.add(ci)

                # Vintage header
                ws.cell(row=row_num, column=1, value=f"Vintage: {vl}").font = bold
                status_cell = ws.cell(row=row_num, column=2, value="FAIL" if has_diff else "PASS")
                status_cell.fill = red_fill if has_diff else green_fill
                status_cell.font = bold
                row_num += 1

                # Left block
                ws.cell(row=row_num, column=1, value=source_left).font = bold
                ws.cell(row=row_num, column=2, value=table_left)
                row_num += 1
                for ci, cn in enumerate(col_names_left):
                    ws.cell(row=row_num, column=ci + 2, value=cn)
                row_num += 1
                for si, (stat_label, stat_key) in enumerate(zip(stat_rows, stat_keys_l)):
                    ws.cell(row=row_num, column=1, value=stat_label).font = bold
                    for ci, comp in enumerate(comps):
                        val = comp.get(stat_key, "")
                        if val is not None and val != "":
                            try:
                                val = float(val)
                            except (ValueError, TypeError):
                                pass
                        cell = ws.cell(row=row_num, column=ci + 2, value=val if val != "" else None)
                        if (ci, si) in diff_cells:
                            cell.fill = red_fill
                        elif ci in problem_cols and si > 0:
                            # Matching stat on a problematic column → green
                            cell.fill = green_fill
                    row_num += 1

                row_num += 1  # blank row

                # Right block
                ws.cell(row=row_num, column=1, value=source_right).font = bold
                ws.cell(row=row_num, column=2, value=table_right)
                row_num += 1
                for ci, cn in enumerate(col_names_right):
                    ws.cell(row=row_num, column=ci + 2, value=cn)
                row_num += 1
                for si, (stat_label, stat_key) in enumerate(zip(stat_rows, stat_keys_r)):
                    ws.cell(row=row_num, column=1, value=stat_label).font = bold
                    for ci, comp in enumerate(comps):
                        val = comp.get(stat_key, "")
                        if val is not None and val != "":
                            try:
                                val = float(val)
                            except (ValueError, TypeError):
                                pass
                        cell = ws.cell(row=row_num, column=ci + 2, value=val if val != "" else None)
                        if (ci, si) in diff_cells:
                            cell.fill = red_fill
                        elif ci in problem_cols and si > 0:
                            cell.fill = green_fill
                    row_num += 1

                row_num += 2  # gap before next vintage

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=col_compare.xlsx"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Row Compare Excel Export (all pairs)
# ---------------------------------------------------------------------------

@app.get("/api/compare/row/export/excel-all")
async def api_compare_row_export_excel_all(from_date: str = "", to_date: str = ""):
    """Generate Excel with one worksheet per pair for row comparison."""
    try:
        from openpyxl import Workbook
        from ..db import list_table_pairs, get_metadata
        from ..compare import compare_row_counts
        from ..config import load_unified_config

        config = load_unified_config(_config_path)
        pairs = list_table_pairs(_db_path)

        from openpyxl.styles import Font, PatternFill
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="CC0000")
        bold = Font(bold=True)

        wb = Workbook()
        wb.remove(wb.active)

        for pair in pairs:
            pname = pair["pair_name"]
            pair_cfg = config.get("pairs", {}).get(pname, {})
            if pair_cfg.get("skip"):
                continue

            comparison = compare_row_counts(
                _db_path, pair["table_left"], pair["table_right"],
                from_date=from_date or None, to_date=to_date or None,
            )

            meta_left = get_metadata(_db_path, pair["table_left"]) or {}
            meta_right = get_metadata(_db_path, pair["table_right"]) or {}
            source_left = (pair.get("source_left") or "LEFT").upper()
            source_right = (pair.get("source_right") or "RIGHT").upper()
            table_left = meta_left.get("source_table") or pair["table_left"]
            table_right = meta_right.get("source_table") or pair["table_right"]
            date_var = meta_left.get("date_var") or meta_right.get("date_var") or ""
            date_fmt = meta_left.get("date_format") or meta_right.get("date_format") or ""

            ws = wb.create_sheet(title=pname[:31])

            # Summary header
            s = comparison["summary"]
            overlap_start = overlap_end = None
            lmin, lmax = s["date_range_left"]
            rmin, rmax = s["date_range_right"]
            if lmin and rmin and lmax and rmax:
                overlap_start = max(lmin, rmin)
                overlap_end = min(lmax, rmax)

            ws.cell(row=1, column=1, value=f"{source_left}:").font = bold
            ws.cell(row=1, column=2, value=table_left)
            ws.cell(row=2, column=1, value=f"{source_right}:").font = bold
            ws.cell(row=2, column=2, value=table_right)
            date_label = f"{date_var} ({date_fmt})" if date_fmt else date_var
            ws.cell(row=3, column=1, value="Date Variable:").font = bold
            ws.cell(row=3, column=2, value=date_label)
            ws.cell(row=4, column=1, value="Overlap:").font = bold
            ws.cell(row=4, column=2, value=f"{overlap_start or '?'} -> {overlap_end or '?'}")
            ws.cell(row=5, column=1, value="Matching:").font = bold
            ws.cell(row=5, column=2, value=f"{s['count_left']} dates" if s.get('count_left') else "0")
            ws.cell(row=6, column=1, value="Mismatch:").font = bold
            ws.cell(row=6, column=2, value=len(comparison["mismatched"]))
            ws.cell(row=7, column=1, value=f"{source_left} Only:").font = bold
            ws.cell(row=7, column=2, value=len(comparison["only_left"]))
            ws.cell(row=8, column=1, value=f"{source_right} Only:").font = bold
            ws.cell(row=8, column=2, value=len(comparison["only_right"]))

            # Date details header
            row_num = 10
            ws.cell(row=row_num, column=1, value="Date").font = bold
            ws.cell(row=row_num, column=2, value=source_left).font = bold
            ws.cell(row=row_num, column=3, value=source_right).font = bold
            ws.cell(row=row_num, column=4, value="Status").font = bold
            row_num += 1

            # Build date rows: mismatches first, then matches
            dates = []
            for dt, l, r in comparison["mismatched"]:
                dates.append((0, dt, l, r, "mismatch"))
            for dt, count in comparison["only_left"]:
                dates.append((0, dt, count, None, f"{source_left.lower()}_only"))
            for dt, count in comparison["only_right"]:
                dates.append((0, dt, None, count, f"{source_right.lower()}_only"))
            for dt, count in comparison["matching"]:
                dates.append((1, dt, count, count, "match"))
            dates.sort(key=lambda r: (r[0], r[1]))

            for sort_key, dt, left, right, status in dates:
                ws.cell(row=row_num, column=1, value=dt)
                ws.cell(row=row_num, column=2, value=left)
                ws.cell(row=row_num, column=3, value=right)
                ws.cell(row=row_num, column=4, value=status)
                if status != "match":
                    for col in range(1, 5):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill = red_fill
                        cell.font = red_font
                row_num += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=row_compare.xlsx"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Row Compare API
# ---------------------------------------------------------------------------

# NOTE: Export routes must be defined BEFORE {pair_name} routes to avoid
# FastAPI matching "export" as a pair_name parameter.

@app.post("/api/compare/row/export/html")
async def api_compare_row_export_html(request: Request):
    """Generate full HTML report for all pairs.

    Body (optional): {
        "from_date": "", "to_date": "",
        "pairs": {
            "pair_name": {
                "excluded_dates": [...],
                "comment_left": "", "comment_right": "",
                "time_left": "", "time_right": ""
            }, ...
        }
    }

    When pairs dict is provided, uses live UI state for exclusions/annotations.
    Otherwise falls back to saved DB/config state.
    """
    body = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    from_date = body.get("from_date", "")
    to_date = body.get("to_date", "")
    title = body.get("title", "") or "Row Count Comparison"
    subtitle = body.get("subtitle", "") or None
    live_pairs = body.get("pairs", {})

    try:
        from ..db import list_table_pairs, get_metadata, get_row_comparison
        from ..compare import compare_row_counts
        from ..html_export import generate_row_count_html, create_row_count_table, wrap_html_document
        from ..config import load_unified_config

        pairs = list_table_pairs(_db_path)
        config = load_unified_config(_config_path)
        sections = []

        for pair in pairs:
            pname = pair["pair_name"]
            pair_cfg = config.get("pairs", {}).get(pname, {})
            if pair_cfg.get("skip"):
                continue

            comparison = compare_row_counts(
                _db_path, pair["table_left"], pair["table_right"],
                from_date=from_date or None, to_date=to_date or None,
            )

            # Use live UI state if provided, otherwise fall back to saved
            live = live_pairs.get(pname, {})
            if live.get("excluded_dates") is not None:
                excluded = set(live["excluded_dates"])
            else:
                saved = get_row_comparison(_db_path, pname)
                excluded = set(saved["excluded_dates"]) if saved and saved.get("excluded_dates") else set()

            if excluded:
                comparison["matching"] = [(dt, c) for dt, c in comparison["matching"] if dt not in excluded]
                comparison["mismatched"] = [(dt, l, r) for dt, l, r in comparison["mismatched"] if dt not in excluded]
                comparison["only_left"] = [(dt, c) for dt, c in comparison["only_left"] if dt not in excluded]
                comparison["only_right"] = [(dt, c) for dt, c in comparison["only_right"] if dt not in excluded]

            meta_left = get_metadata(_db_path, pair["table_left"])
            meta_right = get_metadata(_db_path, pair["table_right"])

            # Annotations: prefer live UI state, fall back to config
            if live:
                comment_left = live.get("comment_left", "")
                comment_right = live.get("comment_right", "")
                time_left = live.get("time_left", "")
                time_right = live.get("time_right", "")
                time_map = {"left": time_left, "right": time_right} if (time_left or time_right) else None
            else:
                comment_map = pair_cfg.get("comment_map", {}).get("row", {})
                comment_left = comment_map.get("left", "")
                comment_right = comment_map.get("right", "")
                time_map = pair_cfg.get("time_map", {}).get("row", {}) or None

            html = generate_row_count_html(
                pair_name=pname,
                source_left=pair.get("source_left", "left"),
                source_right=pair.get("source_right", "right"),
                table_left=pair["table_left"],
                table_right=pair["table_right"],
                comparison=comparison,
                metadata_left=meta_left,
                metadata_right=meta_right,
                comment_left=comment_left,
                comment_right=comment_right,
                time_map=time_map,
                left_cfg=pair_cfg.get("left"),
                right_cfg=pair_cfg.get("right"),
                description=pair_cfg.get("description", ""),
            )
            sections.append(html)

        table = create_row_count_table(sections)
        doc = wrap_html_document(title, [table], subtitle=subtitle)

        return HTMLResponse(content=doc)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/compare/row/export/csv/{pair_name}")
async def api_compare_row_export_csv(pair_name: str, from_date: str = "", to_date: str = ""):
    """Generate CSV export for one pair's row comparison."""
    import csv as csv_mod

    try:
        from ..db import get_table_pair, get_row_comparison
        from ..compare import compare_row_counts

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        comparison = compare_row_counts(
            _db_path, pair["table_left"], pair["table_right"],
            from_date=from_date or None, to_date=to_date or None,
        )

        saved = get_row_comparison(_db_path, pair_name)
        excluded = set(saved["excluded_dates"]) if saved and saved.get("excluded_dates") else set()

        summary = comparison["summary"]
        left_min, left_max = summary["date_range_left"]
        right_min, right_max = summary["date_range_right"]
        overlap_start = overlap_end = None
        if left_min and right_min and left_max and right_max:
            overlap_start = max(left_min, right_min)
            overlap_end = min(left_max, right_max)
            if overlap_start > overlap_end:
                overlap_start = overlap_end = None

        buf = io.StringIO()
        writer = csv_mod.writer(buf)
        writer.writerow(["category", "in_overlap", "excluded", "date", "left_count", "right_count", "diff"])

        rows = []
        for dt, count in comparison["matching"]:
            in_ov = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            rows.append(("match", in_ov, dt in excluded, dt, count, count, 0))
        for dt, l, r in comparison["mismatched"]:
            in_ov = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            rows.append(("mismatch", in_ov, dt in excluded, dt, l, r, r - l))
        for dt, count in comparison["only_left"]:
            in_ov = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            rows.append(("left_only", in_ov, dt in excluded, dt, count, "", ""))
        for dt, count in comparison["only_right"]:
            in_ov = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            rows.append(("right_only", in_ov, dt in excluded, dt, "", count, ""))

        rows.sort(key=lambda r: r[3])
        for row in rows:
            writer.writerow(row)

        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={pair_name}_row_compare.csv"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/compare/row/export/count-csv/{pair_name}")
async def api_compare_row_export_count_csv(
    pair_name: str, side: str = "left", from_date: str = "", to_date: str = "",
    common_only: bool = True,
):
    """Generate a single-side count CSV: DATE_COL (SOURCE), COUNT (SOURCE).

    When common_only=True (default), only dates within the overlap range of
    both sides are included.
    """
    import csv as csv_mod

    try:
        from ..db import get_table_pair, get_metadata
        from ..compare import compare_row_counts

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        if side not in ("left", "right"):
            return JSONResponse(status_code=400, content={"error": "side must be 'left' or 'right'"})

        comparison = compare_row_counts(
            _db_path, pair["table_left"], pair["table_right"],
            from_date=from_date or None, to_date=to_date or None,
        )

        # Compute overlap range
        overlap_start = overlap_end = None
        if common_only:
            summary = comparison["summary"]
            left_min, left_max = summary["date_range_left"]
            right_min, right_max = summary["date_range_right"]
            if left_min and right_min and left_max and right_max:
                overlap_start = max(left_min, right_min)
                overlap_end = min(left_max, right_max)
                if overlap_start > overlap_end:
                    overlap_start = overlap_end = None

        def _in_range(dt):
            if not common_only or not overlap_start:
                return True
            return overlap_start <= dt <= overlap_end

        # Determine column header names
        tbl_key = "table_left" if side == "left" else "table_right"
        src_key = "source_left" if side == "left" else "source_right"
        meta = get_metadata(_db_path, pair[tbl_key]) or {}
        date_var = (meta.get("date_var") or "DT").upper()
        source = (pair.get(src_key) or side).upper()

        # Collect (date, count) rows for the requested side — common range only
        rows = []
        for dt, count in comparison["matching"]:
            if _in_range(dt):
                rows.append((dt, count))
        for dt, l, r in comparison["mismatched"]:
            if _in_range(dt):
                rows.append((dt, l if side == "left" else r))
        if not common_only:
            if side == "left":
                for dt, count in comparison["only_left"]:
                    if _in_range(dt):
                        rows.append((dt, count))
            else:
                for dt, count in comparison["only_right"]:
                    if _in_range(dt):
                        rows.append((dt, count))

        rows.sort(key=lambda r: r[0])

        buf = io.StringIO()
        writer = csv_mod.writer(buf)
        writer.writerow([f"{date_var} ({source})", f"COUNT ({source})"])
        for dt, count in rows:
            writer.writerow([dt, count])

        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={pair_name}_{side}_counts.csv"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/compare/row/export/excel/{pair_name}")
async def api_compare_row_export_excel(
    pair_name: str, from_date: str = "", to_date: str = "",
    common_only: bool = True,
):
    """Generate Excel with two sheets (left counts, right counts) for one pair.

    When common_only=True (default), only dates within the overlap range of
    both sides are included.
    """
    try:
        from openpyxl import Workbook
        from ..db import get_table_pair, get_metadata
        from ..compare import compare_row_counts

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        comparison = compare_row_counts(
            _db_path, pair["table_left"], pair["table_right"],
            from_date=from_date or None, to_date=to_date or None,
        )

        # Compute overlap range
        overlap_start = overlap_end = None
        if common_only:
            summary = comparison["summary"]
            left_min, left_max = summary["date_range_left"]
            right_min, right_max = summary["date_range_right"]
            if left_min and right_min and left_max and right_max:
                overlap_start = max(left_min, right_min)
                overlap_end = min(left_max, right_max)
                if overlap_start > overlap_end:
                    overlap_start = overlap_end = None

        def _in_range(dt):
            if not common_only or not overlap_start:
                return True
            return overlap_start <= dt <= overlap_end

        def _build_side(side):
            tbl_key = "table_left" if side == "left" else "table_right"
            src_key = "source_left" if side == "left" else "source_right"
            meta = get_metadata(_db_path, pair[tbl_key]) or {}
            date_var = (meta.get("date_var") or "DT").upper()
            source = (pair.get(src_key) or side).upper()

            rows = []
            for dt, count in comparison["matching"]:
                if _in_range(dt):
                    rows.append((dt, count))
            for dt, l, r in comparison["mismatched"]:
                if _in_range(dt):
                    rows.append((dt, l if side == "left" else r))
            if not common_only:
                if side == "left":
                    for dt, count in comparison["only_left"]:
                        if _in_range(dt):
                            rows.append((dt, count))
                else:
                    for dt, count in comparison["only_right"]:
                        if _in_range(dt):
                            rows.append((dt, count))
            rows.sort(key=lambda r: r[0])
            return date_var, source, rows

        wb = Workbook()

        for idx, side in enumerate(("left", "right")):
            date_var, source, rows = _build_side(side)
            if idx == 0:
                ws = wb.active
                ws.title = source
            else:
                ws = wb.create_sheet(title=source)
            ws.append([f"{date_var} ({source})", f"COUNT ({source})"])
            for dt, count in rows:
                ws.append([dt, count])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={pair_name}_counts.xlsx"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/compare/row/{pair_name}")
async def api_compare_row_pair(pair_name: str, from_date: str = "", to_date: str = ""):
    """Run row comparison for a single pair and return structured JSON."""
    try:
        from ..db import get_table_pair, get_metadata, get_row_comparison
        from ..compare import compare_row_counts
        from ..config import load_unified_config

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        comparison = compare_row_counts(
            _db_path, pair["table_left"], pair["table_right"],
            from_date=from_date or None, to_date=to_date or None,
        )

        meta_left = get_metadata(_db_path, pair["table_left"]) or {}
        meta_right = get_metadata(_db_path, pair["table_right"]) or {}

        # Load saved comparison state
        saved = get_row_comparison(_db_path, pair_name)

        # Load annotations from config
        config = load_unified_config(_config_path)
        pair_cfg = config.get("pairs", {}).get(pair_name, {})
        comment_map = pair_cfg.get("comment_map", {}).get("row", {})
        time_map = pair_cfg.get("time_map", {}).get("row", {})

        # Build date rows
        summary = comparison["summary"]
        left_min, left_max = summary["date_range_left"]
        right_min, right_max = summary["date_range_right"]

        # Calculate overlap
        overlap_start = overlap_end = None
        if left_min and right_min and left_max and right_max:
            overlap_start = max(left_min, right_min)
            overlap_end = min(left_max, right_max)
            if overlap_start > overlap_end:
                overlap_start = overlap_end = None

        dates = []
        for dt, count in comparison["matching"]:
            in_overlap = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            dates.append({"dt": dt, "left": count, "right": count, "diff": 0, "status": "match", "in_overlap": bool(in_overlap)})
        for dt, l, r in comparison["mismatched"]:
            in_overlap = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            dates.append({"dt": dt, "left": l, "right": r, "diff": r - l, "status": "mismatch", "in_overlap": bool(in_overlap)})
        for dt, count in comparison["only_left"]:
            in_overlap = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            dates.append({"dt": dt, "left": count, "right": None, "diff": None, "status": "left_only", "in_overlap": bool(in_overlap)})
        for dt, count in comparison["only_right"]:
            in_overlap = overlap_start and overlap_end and overlap_start <= dt <= overlap_end
            dates.append({"dt": dt, "left": None, "right": count, "diff": None, "status": "right_only", "in_overlap": bool(in_overlap)})

        dates.sort(key=lambda d: d["dt"] or "")

        return {
            "pair_name": pair_name,
            "table_left": pair["table_left"],
            "table_right": pair["table_right"],
            "source_left": pair.get("source_left", ""),
            "source_right": pair.get("source_right", ""),
            "summary": {
                "count_left": summary["count_left"],
                "count_right": summary["count_right"],
                "total_left": summary["total_left"],
                "total_right": summary["total_right"],
                "date_range_left": list(summary["date_range_left"]),
                "date_range_right": list(summary["date_range_right"]),
                "overlap_start": overlap_start,
                "overlap_end": overlap_end,
                "n_match": len(comparison["matching"]),
                "n_mismatch": len(comparison["mismatched"]),
                "n_left_only": len(comparison["only_left"]),
                "n_right_only": len(comparison["only_right"]),
            },
            "dates": dates,
            "saved": {
                "excluded_dates": saved["excluded_dates"] if saved else [],
                "matching_dates": saved["matching_dates"] if saved else [],
            },
            "annotations": {
                "comment_left": comment_map.get("left", ""),
                "comment_right": comment_map.get("right", ""),
                "time_left": time_map.get("left", ""),
                "time_right": time_map.get("right", ""),
            },
            "meta_left": {
                "date_var": meta_left.get("date_var"),
                "vintage": meta_left.get("vintage"),
            },
            "meta_right": {
                "date_var": meta_right.get("date_var"),
                "vintage": meta_right.get("vintage"),
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/compare/row/{pair_name}")
async def api_save_row_comparison(pair_name: str, request: Request):
    """Save row comparison state: excluded dates, comments, times."""
    body = await request.json()
    try:
        from ..db import save_row_comparison, get_table_pair
        from ..config import load_unified_config, save_unified_config

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        excluded_dates = body.get("excluded_dates", [])
        matching_dates = body.get("matching_dates", [])
        non_matching_dates = body.get("non_matching_dates", [])
        comment_left = body.get("comment_left", "")
        comment_right = body.get("comment_right", "")
        time_left = body.get("time_left", "")
        time_right = body.get("time_right", "")

        # Compute overlap from matching + non-matching dates (full overlap range)
        # Filter out None/empty values that could cause min/max to fail
        all_overlap_dates = [d for d in matching_dates + non_matching_dates if d]
        overlap_start = min(all_overlap_dates) if all_overlap_dates else None
        overlap_end = max(all_overlap_dates) if all_overlap_dates else None

        save_row_comparison(
            _db_path, pair_name,
            overlap_start=overlap_start,
            overlap_end=overlap_end,
            matching_dates=matching_dates,
            excluded_dates=excluded_dates,
            non_matching_dates=non_matching_dates,
        )

        # Save annotations to config
        config = load_unified_config(_config_path)
        pair_cfg = config.get("pairs", {}).get(pair_name, {})
        if pair_cfg:
            pair_cfg.setdefault("comment_map", {})["row"] = {
                "left": comment_left, "right": comment_right,
            }
            pair_cfg.setdefault("time_map", {})["row"] = {
                "left": time_left, "right": time_right,
            }
            save_unified_config(config, _config_path)

        return {"ok": True, "matching": len(matching_dates), "excluded": len(excluded_dates)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/compare/row/{pair_name}/preview")
async def api_compare_row_preview(pair_name: str, request: Request):
    """Generate HTML preview for one pair's row comparison."""
    body = await request.json()
    try:
        from ..db import get_table_pair, get_metadata
        from ..compare import compare_row_counts
        from ..html_export import generate_row_count_html
        from ..config import load_unified_config

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        from_date = body.get("from_date")
        to_date = body.get("to_date")
        excluded_dates = set(body.get("excluded_dates", []))

        comparison = compare_row_counts(
            _db_path, pair["table_left"], pair["table_right"],
            from_date=from_date, to_date=to_date,
        )

        # Filter out excluded dates from all categories
        if excluded_dates:
            comparison["matching"] = [(dt, c) for dt, c in comparison["matching"] if dt not in excluded_dates]
            comparison["mismatched"] = [(dt, l, r) for dt, l, r in comparison["mismatched"] if dt not in excluded_dates]
            comparison["only_left"] = [(dt, c) for dt, c in comparison["only_left"] if dt not in excluded_dates]
            comparison["only_right"] = [(dt, c) for dt, c in comparison["only_right"] if dt not in excluded_dates]

        meta_left = get_metadata(_db_path, pair["table_left"])
        meta_right = get_metadata(_db_path, pair["table_right"])

        config = load_unified_config(_config_path)
        pair_cfg = config.get("pairs", {}).get(pair_name, {})

        time_map = {}
        tm = body.get("time_left") or body.get("time_right")
        if body.get("time_left") or body.get("time_right"):
            time_map = {"left": body.get("time_left", ""), "right": body.get("time_right", "")}

        html = generate_row_count_html(
            pair_name=pair_name,
            source_left=pair.get("source_left", "left"),
            source_right=pair.get("source_right", "right"),
            table_left=pair["table_left"],
            table_right=pair["table_right"],
            comparison=comparison,
            metadata_left=meta_left,
            metadata_right=meta_right,
            comment_left=body.get("comment_left", ""),
            comment_right=body.get("comment_right", ""),
            time_map=time_map or None,
            left_cfg=pair_cfg.get("left"),
            right_cfg=pair_cfg.get("right"),
            description=pair_cfg.get("description", ""),
        )

        # Wrap in a minimal table
        from ..html_export import create_row_count_table
        table_html = create_row_count_table([html])

        return {"html": table_html}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/status/row")
async def api_status_row():
    """Get row count loading status."""
    try:
        from ..db import list_table_pairs, get_metadata

        pairs = list_table_pairs(_db_path)
        status = []

        for pair in pairs:
            for side in ['left', 'right']:
                table = pair[f"table_{side}"]
                meta = get_metadata(_db_path, table) or {}

                status.append({
                    "pair": pair["pair_name"],
                    "side": side,
                    "loaded": meta.get("row_count_total", 0),
                    "min_date": meta.get("min_date_loaded"),
                    "max_date": meta.get("max_date_loaded"),
                })

        return {"status": status, "lastLoad": datetime.now().isoformat()}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/clear/row")
async def api_clear_row():
    """Clear all row count data."""
    try:
        conn = sqlite3.connect(_db_path)
        conn.execute("DELETE FROM _row_counts")
        conn.execute("UPDATE _metadata SET row_count_total = 0, min_date_loaded = NULL, max_date_loaded = NULL")
        conn.commit()
        conn.close()

        return {"ok": True}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Pair management (legacy - keep for compatibility)
# ---------------------------------------------------------------------------

@app.post("/api/pairs")
async def api_add_pair(request: Request):
    """Add a new pair to the config and register in DB.

    Body: {
        "pair_name": "cust_daily",
        "left": {"source": "oracle", "table": "CUST_DAILY", "conn_macro": "pb23", "date_col": "RPT_DT"},
        "right": {"source": "aws", "table": "cust_daily", "conn_macro": "mydb", "date_col": "rpt_dt"},
        "col_map": {}
    }
    """
    body = await request.json()
    pair_name = body.get("pair_name", "").strip()
    if not pair_name:
        return JSONResponse(status_code=400, content={"error": "pair_name is required"})

    left = body.get("left", {})
    right = body.get("right", {})
    if not left.get("table") or not right.get("table"):
        return JSONResponse(status_code=400, content={
            "error": "Both left.table and right.table are required"
        })

    try:
        from ..config import load_unified_config, save_unified_config
        from ..platforms.base import qualified_name
        from ..db import register_table_pair, update_metadata

        config = load_unified_config(_config_path)

        _prepare_side(left)
        _prepare_side(right)

        # Build pair config — only persist valid pair-config keys
        pair_cfg = {
            "left": left,
            "right": right,
            "col_map": body.get("col_map", {}),
        }
        for key in ("description", "mode", "fromDate", "toDate",
                     "dateRangeMode", "overlap"):
            if key in body:
                pair_cfg[key] = body[key]
        config.setdefault("pairs", {})[pair_name] = pair_cfg
        save_unified_config(config, _config_path)

        # Inject name for DB registration (auto-derived, not persisted).
        # Each side derives its own name from `table` so left/right qnames
        # don't collide when both sides share a source.
        from ..config import _derive_side_name
        left["name"] = _derive_side_name(left, pair_name)
        right["name"] = _derive_side_name(right, pair_name)
        table_left = qualified_name(left)
        table_right = qualified_name(right)
        register_table_pair(
            _db_path, pair_name, table_left, table_right,
            source_left=left.get("source"),
            source_right=right.get("source"),
            col_mappings=body.get("col_map") or None,
        )

        # Create metadata stubs so they show up in status
        for tbl_cfg in [left, right]:
            qname = qualified_name(tbl_cfg)
            update_metadata(_db_path, {
                "table_name": qname,
                "source": tbl_cfg.get("source"),
                "source_table": tbl_cfg.get("table"),
                "date_var": tbl_cfg.get("date_col"),
                "data_type": "row",
            })

        return {"ok": True, "pair_name": pair_name,
                "table_left": table_left, "table_right": table_right}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/api/pairs/{pair_name}")
async def api_delete_pair(pair_name: str, purge: int = Query(0)):
    """Remove a pair from config. With purge=1, also remove all DB data."""
    try:
        from ..config import load_unified_config
        config = load_unified_config(_config_path)

        found_in_config = pair_name in config.get("pairs", {})

        if found_in_config:
            del config["pairs"][pair_name]
            with open(_config_path, 'w') as f:
                json.dump(config, f, indent=2)

        found_in_db = False
        if purge:
            from ..db import delete_pair
            try:
                delete_pair(_db_path, pair_name)
                found_in_db = True
            except ValueError:
                pass  # pair may not be registered in DB

        if not found_in_config and not found_in_db:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        return {"ok": True, "deleted": pair_name}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


def _prepare_side(tbl_cfg):
    """Prepare a left/right config for saving.

    - Strips 'name' (auto-derived from pair_name on load)
    - Converts processed string to list for readable JSON
    """
    tbl_cfg.pop("name", None)
    # Store processed as list for readable JSON
    p = tbl_cfg.get("processed", "")
    if isinstance(p, str) and p.strip():
        tbl_cfg["processed"] = [line for line in p.split("\n") if line.strip()]
    elif not p:
        tbl_cfg.pop("processed", None)


def _sync_config_pairs(config):
    """Register all pairs from config into the database.

    Uses patch_metadata for existing rows to avoid wiping data-derived fields.
    Then calls refresh_metadata_from_data to recompute from _row_counts/_col_stats.
    Returns count of pairs synced.
    """
    from ..platforms.base import qualified_name
    from ..db import (
        register_table_pair, update_metadata, patch_metadata,
        get_metadata, refresh_metadata_from_data,
    )

    count = 0
    for pair_name, pair_cfg in config.get("pairs", {}).items():
        if pair_cfg.get("skip"):
            continue
        left = pair_cfg.get("left", {})
        right = pair_cfg.get("right", {})
        if not left.get("source") or not right.get("source"):
            continue

        # Auto-derive name from pair_name (may not be set if config was loaded raw)
        left.setdefault("name", pair_name)
        right.setdefault("name", pair_name)

        table_left = qualified_name(left)
        table_right = qualified_name(right)
        register_table_pair(
            _db_path, pair_name, table_left, table_right,
            source_left=left.get("source"),
            source_right=right.get("source"),
            col_mappings=pair_cfg.get("col_map") or None,
        )

        for tbl_cfg in [left, right]:
            qname = qualified_name(tbl_cfg)
            existing = get_metadata(_db_path, qname)

            # Config-derived fields
            cfg_fields = {
                "source": tbl_cfg.get("source"),
                "source_table": tbl_cfg.get("table"),
                "date_var": tbl_cfg.get("date_col"),
            }
            if tbl_cfg.get("vintage"):
                cfg_fields["vintage"] = tbl_cfg["vintage"]
            # Derive date_format from config date_type
            from ..platforms.base import DATE_TYPE_FORMATS
            cfg_date_type = (tbl_cfg.get("date_type") or "").lower()
            if cfg_date_type in DATE_TYPE_FORMATS:
                cfg_fields["date_format"] = DATE_TYPE_FORMATS[cfg_date_type]

            if existing:
                # Patch without wiping data-derived fields
                patch_metadata(_db_path, qname, **cfg_fields)
            else:
                cfg_fields["table_name"] = qname
                update_metadata(_db_path, cfg_fields)

        count += 1

    # Recompute data-derived fields (min/max dates, row_count_total, date_format)
    # from actual _row_counts and _col_stats data
    refresh_metadata_from_data(_db_path)

    return count


# ---------------------------------------------------------------------------
# HITL annotation routes (where_map, time_map, comment_map)
# ---------------------------------------------------------------------------

@app.get("/api/pairs/{pair_name}/annotations")
async def api_get_annotations(pair_name: str):
    """Get where_map, time_map, comment_map for a pair."""
    try:
        from ..config import load_unified_config
        config = load_unified_config(_config_path)

        pair_cfg = config.get("pairs", {}).get(pair_name)
        if not pair_cfg:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        return {
            "pair_name": pair_name,
            "where_map": pair_cfg.get("where_map", {"left": "", "right": ""}),
            "time_map": pair_cfg.get("time_map", {
                "row": {"left": "", "right": ""},
                "col": {"left": "", "right": ""},
            }),
            "comment_map": pair_cfg.get("comment_map", {
                "row": {"left": "", "right": ""},
                "col": {"left": "", "right": ""},
            }),
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/pairs/{pair_name}/annotations")
async def api_update_annotations(pair_name: str, request: Request):
    """Update where_map, time_map, comment_map for a pair.

    Body (all fields optional):
    {
        "where_map": {"left": "...", "right": "..."},
        "time_map": {"row": {"left": "42s", "right": "112s"}, "col": {...}},
        "comment_map": {"row": {"left": "note", "right": ""}, "col": {...}}
    }
    """
    body = await request.json()
    try:
        from ..config import load_unified_config, save_unified_config
        config = load_unified_config(_config_path)

        if pair_name not in config.get("pairs", {}):
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        pair_cfg = config["pairs"][pair_name]

        if "where_map" in body:
            pair_cfg["where_map"] = body["where_map"]
        if "time_map" in body:
            pair_cfg.setdefault("time_map", {}).update(body["time_map"])
        if "comment_map" in body:
            pair_cfg.setdefault("comment_map", {}).update(body["comment_map"])

        save_unified_config(config, _config_path)

        # Sync where_map to database
        if "where_map" in body:
            from ..db import sync_config_to_db
            sync_config_to_db(_db_path, config)

        return {"ok": True}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Query preview
# ---------------------------------------------------------------------------

@app.post("/api/preview")
async def api_preview(request: Request):
    """Generate preview SQL for a pair's left and right sides.

    Body: {pair_name, left: {...}, right: {...}, fromDate, toDate}
    Returns: {left: {sql, output_file}, right: {sql, output_file}}
    """
    body = await request.json()
    pair_name = body.get("pair_name", "pair")
    from_date = body.get("fromDate", "")
    to_date = body.get("toDate", "")
    njob_left = body.get("njob_left", 0)
    njob_right = body.get("njob_right", 0)

    result = {}
    for side in ("left", "right"):
        cfg = body.get(side, {})
        if not cfg.get("table"):
            continue
        parallel = njob_left if side == "left" else njob_right
        result[side] = _build_preview_sql(
            pair_name, side, cfg, from_date, to_date, parallel=parallel
        )

    return result


def _build_preview_sql(pair_name, side, cfg, from_date, to_date, parallel=0):
    """Build preview SQL for one side of a pair."""
    source = (cfg.get("source") or "").lower()
    date_col = cfg.get("date_col", "")
    date_type = (cfg.get("date_type") or "").lower()
    vintage_raw = cfg.get("vintage", "")
    where_clause = cfg.get("where", "")
    processed = cfg.get("processed", "")
    table = cfg.get("table", "")
    conn_macro = cfg.get("conn_macro", "")

    is_oracle = source == "oracle"
    is_sas = source == "sas"
    is_aws = source == "aws"

    # --- SELECT/GROUP BY expression (with TRUNC/datepart for bucketing) ---
    select_expr = _preview_select_expr(date_col, date_type, source)
    if vintage_raw:
        select_expr = vintage_raw

    # --- WHERE uses raw column (no TRUNC), adjusts literals instead ---
    where_col = date_col

    # --- Resolve table and CTE ---
    cte_prefix = ""
    query_table = table

    if is_sas:
        query_table = f"{conn_macro}.{table}" if conn_macro else table
    elif is_aws:
        if processed:
            alias = f"{pair_name}_{side}"
            cte_prefix = f"WITH {alias} AS (\n  {processed}\n)\n"
            query_table = alias
        else:
            query_table = f"{conn_macro}.{table}" if conn_macro else table
    else:  # oracle
        if processed:
            alias = cfg.get("name", f"{pair_name}_{side}")
            cte_prefix = f"WITH {alias} AS ({processed}) "
            query_table = alias

    # --- WHERE conditions ---
    conditions = []
    _lit = lambda d: _preview_date_literal(d, date_type, is_sas, source)
    if from_date and to_date:
        conditions.append(f"{where_col} BETWEEN {_lit(from_date)} AND {_lit(to_date)}")
    elif from_date:
        conditions.append(f"{where_col} >= {_lit(from_date)}")
    elif to_date:
        conditions.append(f"{where_col} <= {_lit(to_date)}")
    if where_clause:
        conditions.append(f"({where_clause})")

    where_str = ""
    if conditions:
        where_str = "WHERE " + "\n      AND ".join(conditions)

    # --- Parallel hint ---
    parallel_int = int(parallel) if parallel else 0
    ora_hint = f" /*+ PARALLEL({parallel_int}) */" if parallel_int > 1 and (is_oracle or is_sas) else ""
    aws_hint = f"  -- max_workers={parallel_int}" if parallel_int > 1 and is_aws else ""

    # --- Build SQL by platform ---
    output_file = f"{pair_name}_{side}_row"

    if is_aws:
        sql = f"""{cte_prefix}SELECT
  {select_expr} AS date_value,
  COUNT(*) AS row_count
FROM {query_table}
{where_str}
GROUP BY {select_expr}
ORDER BY date_value;{aws_hint}""".strip()
        return {"sql": sql, "output_file": f"{output_file}.sql"}

    elif is_sas:
        # SAS setup + proc sql on local dataset
        setup = ""
        if processed:
            setup = f"{processed}\n\n"
        indent_where = where_str.replace("\n", "\n    ") if where_str else ""
        sql = f"""{setup}proc sql;
  create table work.{output_file} as
  select
    {select_expr} as date_value,
    count(*) as row_count
  from {query_table}
  {indent_where}
  group by {select_expr};
quit;""".strip()
        return {"sql": sql, "output_file": f"{output_file}.sas"}

    else:
        # Oracle via SAS proc sql passthrough
        indent_where = ("    " + where_str.replace("\n", "\n    ")) if where_str else ""
        inner_sql = f"""{cte_prefix}SELECT{ora_hint}
      {select_expr} AS date_value,
      COUNT(*) AS row_count
    FROM {query_table}
    {indent_where}
    GROUP BY {select_expr}"""
        sql = f"""proc sql;
  %{conn_macro}
  create table work.{output_file} as
  select * from connection to oracle (
    {inner_sql.strip()}
  );
  disconnect from oracle;
quit;""".strip()
        return {"sql": sql, "output_file": f"{output_file}.sas"}


def _preview_select_expr(date_col, date_type, source):
    """Build the date expression for SELECT/GROUP BY (bucketing to day level)."""
    dtype = date_type.lower() if date_type else ""

    if source == "oracle":
        if dtype == "timestamp":
            return f"TRUNC({date_col})"
        return date_col
    elif source == "sas":
        if dtype in ("datetime", "timestamp"):
            return f"datepart({date_col})"
        return date_col
    elif source == "aws":
        if dtype in ("timestamp", "datetime"):
            return f"CAST({date_col} AS DATE)"
        return date_col
    return date_col


def _preview_date_literal(date_str, date_type, is_sas=False, source=""):
    """Format a YYYY-MM-DD date string as the correct SQL literal for the date type.

    For timestamp columns, uses TIMESTAMP literals so the raw column can be
    compared without wrapping in TRUNC (preserves index usage).
    """
    if not date_str:
        return "''"
    dtype = date_type.lower() if date_type else ""

    if dtype == "num_yyyymm":
        return date_str.replace("-", "")[:6]

    if dtype in ("num", "integer", "int", "number"):
        return date_str.replace("-", "")

    if dtype == "string_compact":
        return f"'{date_str.replace('-', '')}'"

    if dtype in ("string_dash", "string"):
        return f"'{date_str}'"

    if is_sas and dtype in ("datetime", "date", "timestamp"):
        # SAS date literal: '01JAN2024'd
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(date_str, "%Y-%m-%d")
            if dtype in ("datetime", "timestamp"):
                return f"'{d.strftime('%d%b%Y').upper()}:00:00:00'dt"
            return f"'{d.strftime('%d%b%Y').upper()}'d"
        except ValueError:
            return f"'{date_str}'"

    # Oracle/Athena timestamp: use TIMESTAMP literal
    if dtype == "timestamp" and source == "oracle":
        return f"TIMESTAMP '{date_str} 00:00:00'"
    if dtype in ("timestamp", "datetime") and source == "aws":
        return f"TIMESTAMP '{date_str} 00:00:00'"

    # Oracle/Athena date: DATE literal
    if dtype in ("date",):
        return f"DATE '{date_str}'"

    return f"'{date_str}'"


# ---------------------------------------------------------------------------
# Column Mapping API
# ---------------------------------------------------------------------------

@app.post("/api/pairs/{pair_name}/col_filter/preview")
async def api_col_filter_preview(pair_name: str, request: Request):
    """Resolve include/exclude patterns against the pair's col_map and return
    the effective (left → right) pairs plus a warning for patterns that hit
    left columns not in col_map.

    Body: {"include": ["CUST_*", ...], "exclude": ["*_AUDIT*", ...]}
    """
    from fnmatch import fnmatch

    body = await request.json()
    include = [p.strip() for p in (body.get("include") or []) if p and p.strip()]
    exclude = [p.strip() for p in (body.get("exclude") or []) if p and p.strip()]
    try:
        from ..config import load_unified_config
        from ..db import get_table_pair, get_column_meta, get_pair_col_map_from_db
        from ..compare import resolve_col_filter

        config = load_unified_config(_config_path)
        pair_cfg = config.get("pairs", {}).get(pair_name)
        if not pair_cfg:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        # Prefer DB-stored col_map (col_mapping page writes there), fall back
        # to the one in the JSON config if it exists.
        col_map = get_pair_col_map_from_db(_db_path, pair_name) or pair_cfg.get("col_map") or {}

        resolved = resolve_col_filter(col_map, include, exclude)
        pairs = resolved["pairs"]

        # Patterns can also name left cols that aren't in col_map — e.g. the
        # user picked a column that hasn't been mapped. Surface that so the
        # card can warn explicitly.
        unmapped_matches = []
        if include:
            pair = get_table_pair(_db_path, pair_name)
            left_meta = get_column_meta(_db_path, pair["table_left"]) if pair else []
            mapped_left_lower = {l.lower() for l in col_map}
            include_lower = [p.lower() for p in include]
            for cm in left_meta:
                cn = cm["column_name"]
                cnl = cn.lower()
                if cnl in mapped_left_lower:
                    continue
                if any(fnmatch(cnl, p) for p in include_lower):
                    # But if also in exclude, skip
                    if any(fnmatch(cnl, p.lower()) for p in exclude):
                        continue
                    unmapped_matches.append(cn)

        return {
            "ok": True,
            "total_mapped": resolved["total_mapped"],
            "effective_count": resolved["effective_count"],
            "pairs": [{"left": l, "right": r} for l, r in pairs],
            "unmapped_matches": sorted(unmapped_matches, key=str.lower),
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/pairs/{pair_name}/columns")
async def api_get_pair_columns(pair_name: str):
    """Return left/right columns, current col_mappings, col_rules, and auto-match results."""
    try:
        from ..db import (
            get_table_pair, get_column_meta, get_pair_col_map_from_db,
            get_pair_col_rules, ensure_col_rules_column,
        )
        from ..compare import match_columns_from_dicts

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        ensure_col_rules_column(_db_path)

        # Get column metadata for both sides
        left_meta = get_column_meta(_db_path, pair["table_left"])
        right_meta = get_column_meta(_db_path, pair["table_right"])

        left_cols = {c["column_name"]: c.get("data_type", "") for c in left_meta}
        right_cols = {c["column_name"]: c.get("data_type", "") for c in right_meta}

        # Get existing mappings and rules
        col_mappings = get_pair_col_map_from_db(_db_path, pair_name)
        col_rules_data = get_pair_col_rules(_db_path, pair_name)

        # Determine unmapped columns
        mapped_left = set(col_mappings.keys())
        mapped_right = set(col_mappings.values())
        unmapped_left = {k: v for k, v in left_cols.items() if k not in mapped_left}
        unmapped_right = {k: v for k, v in right_cols.items() if k not in mapped_right}

        # Auto-match unmapped columns
        auto_result = match_columns_from_dicts(
            unmapped_left, unmapped_right,
            left_label=pair.get("source_left", "left"),
            right_label=pair.get("source_right", "right"),
        )

        return {
            "pair_name": pair_name,
            "table_left": pair["table_left"],
            "table_right": pair["table_right"],
            "source_left": pair.get("source_left", ""),
            "source_right": pair.get("source_right", ""),
            "left_columns": left_cols,
            "right_columns": right_cols,
            "col_mappings": col_mappings,
            "col_rules": col_rules_data,
            "auto_match": auto_result.get("matched", {}),
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/pairs/{pair_name}/columns/csv")
async def api_get_pair_columns_csv(pair_name: str, side: str = "left"):
    """Generate a column metadata CSV for one side of a pair.

    Columns: TABLE_NAME, COLUMN_NAME, DATA_TYPE, DATA_TYPE_EXTENDED, MANDATORY
    """
    import csv as csv_mod

    try:
        from ..db import get_table_pair, get_column_meta

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        if side not in ("left", "right"):
            return JSONResponse(status_code=400, content={"error": "side must be 'left' or 'right'"})

        tbl_key = "table_left" if side == "left" else "table_right"
        table_name = pair[tbl_key].upper()
        col_meta = get_column_meta(_db_path, pair[tbl_key])

        buf = io.StringIO()
        writer = csv_mod.writer(buf)
        writer.writerow(["TABLE_NAME", "COLUMN_NAME", "DATA_TYPE", "DATA_TYPE_EXTENDED", "MANDATORY"])

        for col in sorted(col_meta, key=lambda c: c["column_name"]):
            dt = (col.get("data_type") or "").upper()
            writer.writerow([
                table_name,
                col["column_name"].upper(),
                dt,
                "",
                "",
            ])

        buf.seek(0)
        src_key = "source_left" if side == "left" else "source_right"
        source = pair.get(src_key, side)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={pair_name}_{source}_columns.csv"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/pairs/{pair_name}/columns/excel")
async def api_get_pair_columns_excel(pair_name: str):
    """Generate Excel with two sheets (left columns, right columns) for one pair."""
    try:
        from openpyxl import Workbook
        from ..db import get_table_pair, get_column_meta

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        wb = Workbook()
        for idx, side in enumerate(("left", "right")):
            tbl_key = "table_left" if side == "left" else "table_right"
            src_key = "source_left" if side == "left" else "source_right"
            table_name = pair[tbl_key].upper()
            source = (pair.get(src_key) or side).upper()
            col_meta = get_column_meta(_db_path, pair[tbl_key])

            if idx == 0:
                ws = wb.active
                ws.title = source
            else:
                ws = wb.create_sheet(title=source)

            ws.append(["TABLE_NAME", "COLUMN_NAME", "DATA_TYPE", "DATA_TYPE_EXTENDED", "MANDATORY"])
            for col in sorted(col_meta, key=lambda c: c["column_name"]):
                dt = (col.get("data_type") or "").upper()
                ws.append([table_name, col["column_name"].upper(), dt, "", ""])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={pair_name}_columns.xlsx"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/pairs/{pair_name}/col-type-overrides")
async def api_get_col_type_overrides(pair_name: str):
    """Get column type overrides for a pair."""
    try:
        from ..config import load_unified_config, get_col_type_overrides
        config = load_unified_config(_config_path)
        overrides = get_col_type_overrides(config, pair_name)
        return {"ok": True, "pair_name": pair_name, "overrides": overrides}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/pairs/{pair_name}/col-type-overrides")
async def api_save_col_type_overrides(pair_name: str, request: Request):
    """Save column type overrides for a pair.

    Body: {overrides: {col_name: "numeric"|"categorical", ...}}
    """
    body = await request.json()
    try:
        from ..config import load_unified_config, save_unified_config, set_col_type_override

        config = load_unified_config(_config_path)
        if pair_name not in config.get("pairs", {}):
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        overrides = body.get("overrides", {})
        # Replace the entire overrides dict
        config["pairs"][pair_name]["col_type_overrides"] = overrides
        save_unified_config(config, _config_path)

        return {"ok": True, "pair_name": pair_name, "saved": len(overrides)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/api/pairs/{pair_name}/col-mappings")
async def api_save_col_mappings(pair_name: str, request: Request):
    """Save flat mappings + rules + sources."""
    body = await request.json()
    try:
        from ..db import (
            get_table_pair, update_pair_col_map,
            update_pair_col_rules, ensure_col_rules_column,
        )

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        ensure_col_rules_column(_db_path)

        mappings = body.get("mappings", {})
        rules_data = body.get("rules", {})

        update_pair_col_map(_db_path, pair_name, mappings)
        update_pair_col_rules(_db_path, pair_name, rules_data)

        return {"ok": True, "mapped": len(mappings)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/pairs/{pair_name}/col-mappings/csv")
async def api_col_mappings_csv(pair_name: str):
    """Download CSV with columns: status, left_column, left_type, right_column, right_type."""
    import csv as csv_mod

    try:
        from ..db import get_table_pair, get_column_meta, get_pair_col_map_from_db

        pair = get_table_pair(_db_path, pair_name)
        if not pair:
            return JSONResponse(status_code=404, content={"error": f"Pair '{pair_name}' not found"})

        left_meta = get_column_meta(_db_path, pair["table_left"])
        right_meta = get_column_meta(_db_path, pair["table_right"])

        left_types = {c["column_name"]: c.get("data_type", "") for c in left_meta}
        right_types = {c["column_name"]: c.get("data_type", "") for c in right_meta}

        col_mappings = get_pair_col_map_from_db(_db_path, pair_name)
        source_left = pair.get("source_left", "left")
        source_right = pair.get("source_right", "right")

        buf = io.StringIO()
        writer = csv_mod.writer(buf)
        writer.writerow(["status", "left_column", "left_type", "right_column", "right_type"])

        # Mapped columns
        for left_col, right_col in sorted(col_mappings.items()):
            writer.writerow([
                "matched",
                left_col,
                left_types.get(left_col, ""),
                right_col,
                right_types.get(right_col, ""),
            ])

        # Left-only
        mapped_left = set(col_mappings.keys())
        for col in sorted(left_types.keys()):
            if col not in mapped_left:
                writer.writerow([
                    f"{source_left}-only",
                    col,
                    left_types.get(col, ""),
                    "",
                    "",
                ])

        # Right-only
        mapped_right = set(col_mappings.values())
        for col in sorted(right_types.keys()):
            if col not in mapped_right:
                writer.writerow([
                    f"{source_right}-only",
                    "",
                    "",
                    col,
                    right_types.get(col, ""),
                ])

        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={pair_name}_col_mappings.csv"},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# ---------------------------------------------------------------------------
# Testing mode
# ---------------------------------------------------------------------------

@app.get("/api/testing")
async def api_testing_status():
    """Return current testing mode state."""
    return {"testing": _testing_mode, "config_path": _config_path}


@app.post("/api/testing")
async def api_testing_toggle(request: Request):
    """Toggle testing mode on/off.

    When enabled: sets DTRACK_MOCK env var, swaps config to testing/config.json.
    When disabled: restores original config, unsets DTRACK_MOCK.
    """
    global _config_path, _original_config_path, _testing_mode

    body = await request.json()
    enable = body.get("enabled", False)

    testing_dir = Path(__file__).parent.parent.parent / "testing"
    mock_dir = testing_dir / "mock"
    test_config = testing_dir / "config.json"

    if enable:
        if not test_config.exists():
            return JSONResponse(status_code=404, content={
                "error": f"Testing config not found: {test_config}"
            })
        if not mock_dir.exists():
            return JSONResponse(status_code=404, content={
                "error": f"Mock data not found: {mock_dir}"
            })
        _original_config_path = _config_path
        _config_path = str(test_config.resolve())
        os.environ["DTRACK_MOCK"] = str(mock_dir.resolve())
        _testing_mode = True
        return {"ok": True, "testing": True,
                "config_path": _config_path,
                "mock_dir": str(mock_dir.resolve())}
    else:
        if _original_config_path:
            _config_path = _original_config_path
            _original_config_path = ""
        os.environ.pop("DTRACK_MOCK", None)
        _testing_mode = False
        return {"ok": True, "testing": False, "config_path": _config_path}


# ---------------------------------------------------------------------------
# Playground: ad-hoc SAS / Athena / Oracle queries with timing + history
# ---------------------------------------------------------------------------

def _row_to_strs(row, n_cols):
    """Coerce a DB row into a list of n_cols string values."""
    out = []
    for i in range(n_cols):
        v = row[i] if i < len(row) else None
        out.append('' if v is None else str(v))
    return out


@app.post("/api/playground/run")
async def api_playground_run(request: Request):
    """Run an ad-hoc SQL query against Athena or Oracle and time it.

    Body: {engine: 'athena'|'oracle', sql, conn?, row_cap?}
    Returns the first `row_cap` rows; the full count is in n_rows_total.
    Always inserts a row into _playground_history (status ok|error).
    """
    import time
    from ..db import insert_playground_run

    body = await request.json()
    engine = (body.get("engine") or "").lower()
    sql = (body.get("sql") or "").strip()
    conn_id = body.get("conn") or ""
    row_cap = int(body.get("row_cap") or 200)

    if engine not in ("athena", "oracle"):
        return JSONResponse(status_code=400, content={"error": f"Unsupported engine: {engine!r}"})
    if not sql:
        return JSONResponse(status_code=400, content={"error": "SQL is required"})

    columns: list = []
    rows: list = []
    n_rows_total = 0
    elapsed = 0.0
    err: Optional[str] = None

    t0 = time.perf_counter()
    try:
        if engine == "athena":
            from ..platforms.athena import athena_connect
            conn = athena_connect(conn_id or None)
            cur = conn.cursor()
            cur.execute(sql)
            columns = [d[0] for d in (cur.description or [])]
            fetched = cur.fetchall() or []
            n_rows_total = len(fetched)
            rows = [_row_to_strs(r, len(columns)) for r in fetched[:row_cap]]
        else:
            from ..db import oracle_connect
            if not conn_id:
                raise ValueError("Oracle requires a connection macro (conn).")
            conn = oracle_connect(conn_id)
            if conn is None:
                raise RuntimeError("oracle_connect returned None (mock mode?)")
            cur = conn.cursor()
            cur.execute(sql)
            columns = [d[0] for d in (cur.description or [])]
            fetched = cur.fetchall() or []
            n_rows_total = len(fetched)
            rows = [_row_to_strs(r, len(columns)) for r in fetched[:row_cap]]
            try:
                cur.close()
                conn.close()
            except Exception:
                pass
    except Exception as e:
        err = str(e)
    elapsed = round(time.perf_counter() - t0, 4)

    # Optional from typing is already imported at module top. We use it via Optional[str] above.
    history_id = insert_playground_run(
        _db_path,
        engine=engine,
        conn=conn_id,
        sql=sql,
        elapsed_sec=elapsed,
        n_rows=n_rows_total if err is None else None,
        status='ok' if err is None else 'error',
        error_msg=err,
    )

    if err is not None:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": err, "elapsed_sec": elapsed,
            "history_id": history_id,
        })

    return {
        "ok": True,
        "columns": columns,
        "rows": rows,
        "n_rows_total": n_rows_total,
        "n_rows_returned": len(rows),
        "elapsed_sec": elapsed,
        "history_id": history_id,
    }


@app.post("/api/playground/sas")
async def api_playground_sas(request: Request):
    """Generate SAS extraction script(s) from the current config.

    Body: {type?: 'row'|'col'|'both', from_date?, to_date?, vintage?}
    Returns: {filename, content} — the user downloads as <filename>.sas.
    Also records a history row (engine='sas', elapsed=None).
    """
    import tempfile
    from ..platforms.oracle import gen_sas
    from ..db import insert_playground_run

    body = await request.json()
    extract_type = (body.get("type") or "both").lower()
    from_date = body.get("from_date") or None
    to_date = body.get("to_date") or None
    vintage = body.get("vintage") or None

    types = ["row", "col"] if extract_type == "both" else [extract_type]

    err: Optional[str] = None
    parts: list = []
    try:
        with tempfile.TemporaryDirectory() as td:
            buf = io.StringIO()
            with redirect_stdout(buf), redirect_stderr(buf):
                gen_sas(_config_path, td, types=types, db_path=_db_path,
                        from_date=from_date, to_date=to_date, vintage=vintage)
            for fname in sorted(os.listdir(td)):
                fpath = os.path.join(td, fname)
                if os.path.isfile(fpath) and fname.endswith('.sas'):
                    with open(fpath, encoding='utf-8') as f:
                        parts.append((fname, f.read()))
    except Exception as e:
        err = str(e)

    history_id = insert_playground_run(
        _db_path,
        engine='sas',
        conn=extract_type,
        sql=f"gen_sas type={extract_type}",
        elapsed_sec=None,
        n_rows=None,
        status='ok' if err is None else 'error',
        error_msg=err,
    )

    if err:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": err, "history_id": history_id,
        })

    if not parts:
        return JSONResponse(status_code=500, content={
            "ok": False, "error": "gen_sas produced no .sas files",
            "history_id": history_id,
        })

    if len(parts) == 1:
        filename, content = parts[0]
    else:
        filename = f"playground_{extract_type}.sas"
        content = "\n\n/* ============================================ */\n\n".join(
            f"/* === {n} === */\n{c}" for n, c in parts
        )

    return {"ok": True, "filename": filename, "content": content,
            "history_id": history_id}


@app.get("/api/playground/history")
async def api_playground_history():
    from ..db import list_playground_runs
    return {"runs": list_playground_runs(_db_path, limit=200)}


@app.put("/api/playground/history/{run_id}")
async def api_playground_history_update(run_id: int, request: Request):
    from ..db import update_playground_note
    body = await request.json()
    note = body.get("note", "")
    ok = update_playground_note(_db_path, run_id, note)
    if not ok:
        return JSONResponse(status_code=404, content={"error": "run not found"})
    return {"ok": True}


@app.delete("/api/playground/history/{run_id}")
async def api_playground_history_delete(run_id: int):
    from ..db import delete_playground_run
    ok = delete_playground_run(_db_path, run_id)
    if not ok:
        return JSONResponse(status_code=404, content={"error": "run not found"})
    return {"ok": True}


# ---------------------------------------------------------------------------
# CSV Compare: string-exact diff between two uploaded CSVs by primary key
# ---------------------------------------------------------------------------

@app.post("/api/csv_compare/inspect")
async def api_csv_compare_inspect(left: UploadFile = File(...),
                                  right: UploadFile = File(...)):
    """Read column lists + row counts from two uploaded CSVs."""
    from ..csv_compare import read_csv_as_str

    try:
        left_df  = read_csv_as_str(io.BytesIO(await left.read()))
        right_df = read_csv_as_str(io.BytesIO(await right.read()))
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": f"CSV parse failed: {e}"})

    return {
        "left":  {"filename": left.filename,  "columns": list(left_df.columns),  "n_rows": int(len(left_df))},
        "right": {"filename": right.filename, "columns": list(right_df.columns), "n_rows": int(len(right_df))},
    }


@app.post("/api/csv_compare/run")
async def api_csv_compare_run(
    left: UploadFile = File(...),
    right: UploadFile = File(...),
    pk_cols: str = Form(...),
    compare_cols: str = Form(...),
    n_examples: int = Form(10),
):
    """Compare two uploaded CSVs by primary key, string-exact, return diffs."""
    from ..csv_compare import read_csv_as_str, compare_csvs

    pk_list = [c.strip() for c in pk_cols.split(',') if c.strip()]
    compare_list = [c.strip() for c in compare_cols.split(',') if c.strip()]
    if not pk_list:
        return JSONResponse(status_code=400, content={"error": "pk_cols is required."})
    if not compare_list:
        return JSONResponse(status_code=400, content={"error": "compare_cols is required."})

    try:
        left_df  = read_csv_as_str(io.BytesIO(await left.read()))
        right_df = read_csv_as_str(io.BytesIO(await right.read()))
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": f"CSV parse failed: {e}"})

    try:
        result = compare_csvs(left_df, right_df, pk_list, compare_list,
                              n_examples=int(n_examples))
    except ValueError as e:
        return JSONResponse(status_code=400, content={"error": str(e)})
    return result


# ---------------------------------------------------------------------------
# Server entry point
# ---------------------------------------------------------------------------

def serve(db_path: str, config_path: str, port: int = 8080, host: str = "0.0.0.0"):
    """Start the dtrack web server."""
    global _db_path, _db_dir, _config_path, _original_config_path
    _db_path = os.path.abspath(db_path)
    _db_dir = os.path.dirname(_db_path)
    _config_path = os.path.abspath(config_path)
    _original_config_path = _config_path

    # Sync config pairs into DB on startup
    try:
        from ..config import load_unified_config
        config = load_unified_config(_config_path)
        n = _sync_config_pairs(config)
        print(f"  Synced {n} pairs from config → DB")
    except Exception as e:
        print(f"  Warning: could not sync pairs: {e}")

    import uvicorn
    print(f"dtrack web UI")
    print(f"  Database:  {_db_path}")
    print(f"  Config:    {_config_path}")
    print(f"  Base dir:  {_db_dir}")
    print(f"  URL:       http://localhost:{port}")
    uvicorn.run(app, host=host, port=port, log_level="warning")
